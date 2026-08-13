#!/usr/bin/env python3
"""Build the isolated complete teacher-vocabulary preview for Issue #185.

The workbook and the teacher image package are explicit local inputs. Their
absolute paths are never copied to repository output. This script is the only
writer for the preview corpus, reconciliation manifests, and deterministic
preview derivatives; it never changes the source inputs or production assets.

Usage:
  uv run --locked python scripts/build-teacher-vocabulary-complete-preview.py \
    --workbook /path/to/单词表(带图).xlsx --source-dir /path/to/词汇表 \
    --inventory /tmp/chabiko_teacher_image_inventory.json --build

Place accepted built-in image-generation outputs at <ai-source-dir>/<preview-id>
with a .png, .webp, or .jpg suffix, then rerun --build. The source files are
only read. Generated preview derivatives are written as deterministic WebP.

Teacher-derived derivatives are written to the tracked public assets path
(--teacher-asset-dir, default public/assets/vocabulary/teacher-preview/teacher)
and recorded in the corpus as teacher-mapped, so they are included in the
deployed static build for remote teacher review. AI-generated preview
derivatives remain under the tracked public AI assets path.

Accepted AI assets are normally reused without rewriting by passing
--reuse-accepted-ai-assets, which verifies each committed ai-generated record
against its on-disk WebP (preview ID, existence, WebP readability and
dimensions, SHA-256, transparent corners, one-to-one mapping) and fails closed
on any drift. Fresh AI inputs continue to use --ai-source-dir. The canonical
teacher-image inventory is produced by scripts/build-teacher-image-inventory.py.

Integrity invariant: an existing teacher derivative is reused only when its
bytes match a deterministic re-export of the reconciliation-verified source
PNG. A changed source, altered/stale derivative, or a derivative copied from
another preview item therefore fails closed and is regenerated, keeping the
fresh-export and reuse paths under the same source checksum/dimension and
derivative byte-integrity gates.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from opencc import OpenCC
from PIL import Image, ImageOps


REPO_ROOT = Path(__file__).resolve().parent.parent
EXPECTED_WORKBOOK_SHA256 = "3fad65934dd3801fedfbd9e110f2c5bb8730b36d4117ee7a228cbf0089383f37"
EXPECTED_IMAGE_FINGERPRINT = "592cee9f32419b4b3571146d72c8710cd9f1edc5e02150c48178011a5a8b1517"
EXPECTED_IMAGE_COUNT = 1240

SHEET_META = {
    "名词1": {"partOfSpeech": "noun"},
    "动词1": {"partOfSpeech": "verb"},
    "形容词1": {"partOfSpeech": "adjective"},
    "副词": {"partOfSpeech": "adverb"},
    "名词2": {"partOfSpeech": "noun"},
    "形容词2": {"partOfSpeech": "adjective"},
    "动词2": {"partOfSpeech": "verb"},
}
SHEET_ORDER = tuple(SHEET_META)
HEADERS = ("单词", "拼音", "日语翻译", "难易度", "造词/造句")

# These phrases cannot safely be resolved to a single concept without adding
# text, context, or an invented interpretation. The list is intentionally
# conservative and applies only after deterministic teacher-image matching.
ABSTRACT_NOUN_MARKERS = frozenset(
    "主义 关系 方法 情况 问题 原因 结果 意义 作用 制度 政治 经济 文化 教育 法律 社会 历史 技术 科学 "
    "经验 知识 能力 机会 条件 特点 目的 计划 决定 希望 意见 要求 消息 新闻 事情 活动 运动 系统 水平 程度 "
    "数量 质量 价格 服务 工作 职业 行业 单位 组织 国家 城市 地区 地点 空间 时间 方向 过程 办法 方面 资料 信息 "
    "变化 影响 责任 权利 义务 任务 习惯 理想 目标 态度 选择 规则 秩序 规定 内容 现象 理论 观念 精神 性格 性质 "
    "价值 未来 过去 现在 可能 性别 年纪 生活 生命 人生 时代 年代"
    .split()
)
TEXT_ONLY_TERMS = frozenset(
    "我 我们 你 你们 他 她 它 他们 她们 它们 咱们 自己 大家 这 那 这里 那里 哪儿 谁 什么 怎么 为什么 "
    "今天 明天 后天 昨天 前天 半天 每天 整天 每周 每年 这周 下周 上周 今年 明年 后年 去年 前年 "
    "早上 晚上 上午 中午 下午 白天 傍晚 以前 以后 将来 最近"
    .split()
)
FUNCTIONAL_VERBS = frozenset(
    "是 有 在 会 要 能 可以 应该 觉得 认为 知道 了解 希望 喜欢 爱 怕 想 需要 属于 成为 存在 发生 进行 "
    "继续 开始 结束 决定 同意 相信 注意 担心 忘记 记得 明白 认识 以为 让 使 给 被 把 叫 说 讲 问 答 必须 愿意 打算 计划 选择 比较 像"
    .split()
)
VISUAL_ADJECTIVES = frozenset(
    "高 低 长 短 大 小 胖 瘦 快 慢 热 冷 白 黑 红 黄 绿 蓝 亮 暗 干 湿 重 轻 好 坏 漂亮 美丽 可爱 安静 吵 "
    "干净 脏 新 旧 满 空 忙 累 饿 渴 甜 酸 辣 苦 咸 圆 方 直 弯 近 远 贵 便宜 年轻 老 健康 生病 危险 安全 "
    "方便 难 容易 简单 复杂 舒服 难受 晴 阴 热闹 拥挤 清楚 模糊 深 浅 厚 薄 宽 窄 高兴 难过 生气 害怕 紧张 放心 温暖 凉快 新鲜 香 臭"
    .split()
)
SAFETY_SENSITIVE_TERMS = frozenset({"脱"})

# The first small built-in generation checkpoint was reviewed before the
# revision-two refinement below. Keep that provenance exact for those accepted
# inputs; later queue entries use the frozen revision-two wording.
CHECKPOINT_PREVIEW_IDS = frozenset({
    "teacher-preview-3beb1fd09cf9f6cd",
    "teacher-preview-e431cd2fee25738f",
    "teacher-preview-8c583020eaacad93",
    "teacher-preview-d246af0696d938f8",
    "teacher-preview-3e18cf3ce51c2daa",
})
STYLE_REFERENCE_SET_IDS = (
    "名词1 1-51/大家.png",
    "人.png",
    "名词1 216-265/水果.png",
    "動詞1-49/看.png",
    "形容詞1-40/好.png",
)
# Items whose accepted revision is not their first generated revision. The
# value is a machine-readable record of the prior rejection under the frozen
# style-audit consistency procedure; the revision-three wording exists solely
# to fix the framing/chroma-key defects named in the reason.
REGENERATION_REASONS: dict[str, dict[str, str]] = {
    "teacher-preview-42233332a5fffab8": {
        "fromRevision": "2",
        "toRevision": "3",
        "reason": "style-consistency-rejection: revision 2 placed the subject outside the central 70% of the canvas and left a border not entirely #00ff00",
        "outcome": "rejected-and-regenerated",
        "evidence": "revision-3 wording adds central-70% framing and full #00ff00 border requirements",
    },
    "teacher-preview-c1b3a4997c033074": {
        "fromRevision": "2",
        "toRevision": "3",
        "reason": "style-consistency-rejection: revision 2 placed the subject outside the central 70% of the canvas and left a border not entirely #00ff00",
        "outcome": "rejected-and-regenerated",
        "evidence": "revision-3 wording adds central-70% framing and full #00ff00 border requirements",
    },
}
REGENERATED_PREVIEW_IDS = frozenset(REGENERATION_REASONS)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize(value: Any) -> str:
    value = "" if value is None else str(value)
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", value).strip())


def stable_preview_id(sheet: str, row: int, word: str) -> str:
    seed = f"teacher-preview-v1|{sheet}|{row}|{normalize(word)}"
    return f"teacher-preview-{hashlib.sha256(seed.encode('utf-8')).hexdigest()[:16]}"


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    raw = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(raw, encoding="utf-8")
    temp.replace(path)


def load_inventory(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    inventory = json.loads(path.read_text(encoding="utf-8"))
    if inventory.get("fingerprint_withpaths") != EXPECTED_IMAGE_FINGERPRINT:
        raise ValueError("Teacher-image package path-sensitive fingerprint does not match #185")
    readable = [record for record in inventory.get("records", []) if record.get("readable")]
    if len(readable) != EXPECTED_IMAGE_COUNT:
        raise ValueError(f"Expected {EXPECTED_IMAGE_COUNT} readable teacher images, found {len(readable)}")
    if any(record.get("fmt") != "PNG" for record in readable):
        raise ValueError("Teacher-image inventory contains a non-PNG readable image")
    return sorted(readable, key=lambda record: record["rel"]), inventory


def load_workbook_rows(workbook_path: Path) -> list[dict[str, Any]]:
    if sha256_file(workbook_path) != EXPECTED_WORKBOOK_SHA256:
        raise ValueError("Canonical workbook SHA-256 does not match #185")
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    rows: list[dict[str, Any]] = []
    for sheet_name in SHEET_ORDER:
        worksheet = workbook[sheet_name]
        header_columns = {
            normalize(worksheet.cell(1, column).value): column
            for column in range(1, worksheet.max_column + 1)
            if normalize(worksheet.cell(1, column).value)
        }
        missing = set(HEADERS) - set(header_columns)
        if missing:
            raise ValueError(f"{sheet_name}: missing workbook headers: {sorted(missing)}")
        for row_number in range(2, worksheet.max_row + 1):
            word = normalize(worksheet.cell(row_number, header_columns["单词"]).value)
            if not word:
                continue
            row_dict = {
                "key": f"{sheet_name}:{row_number}",
                "sourceSheet": sheet_name,
                "sourceRow": row_number,
                "simplified": word,
                "pinyin": normalize(worksheet.cell(row_number, header_columns["拼音"]).value),
                "japanese": normalize(worksheet.cell(row_number, header_columns["日语翻译"]).value),
                "difficulty": normalize(worksheet.cell(row_number, header_columns["难易度"]).value),
                "partOfSpeech": SHEET_META[sheet_name]["partOfSpeech"],
            }
            # Teacher-authored example sentence (#340): preserved verbatim, only
            # normalized. An empty cell is a supported missing-example state, so
            # the field is omitted rather than generated.
            example = normalize(worksheet.cell(row_number, header_columns["造词/造句"]).value)
            if example:
                row_dict["example"] = example
            rows.append(row_dict)
    if len(rows) != 1865:
        raise ValueError(f"Expected 1,865 workbook candidate rows, found {len(rows)}")
    return rows


def production_records() -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    vocabulary = json.loads(
        (REPO_ROOT / "data/vocabulary/teacher-core-v1/teacher-vocabulary-batch-01.json").read_text(encoding="utf-8")
    )["vocabulary"]
    illustrations = json.loads(
        (REPO_ROOT / "data/illustrations/teacher-core-v1/teacher-vocabulary-batch-01.json").read_text(encoding="utf-8")
    )["illustrations"]
    if len(vocabulary) != 20 or len(illustrations) != 19:
        raise ValueError("Production teacher baseline is not the immutable 20-row/19-image contract")
    by_source: dict[str, dict[str, Any]] = {}
    for record in vocabulary:
        curriculum = record["curriculum"]
        by_source[f"{curriculum['sourceSheet']}:{curriculum['sourceRow']}"] = record
    ids_by_checksum: dict[str, list[str]] = defaultdict(list)
    for illustration in illustrations:
        ids_by_checksum[illustration["sourceChecksumSha256"]].append(illustration["vocabularyId"])
    return by_source, ids_by_checksum


def source_duplicate_metadata(inventory: dict[str, Any]) -> tuple[dict[str, str], dict[str, str]]:
    group_for_path: dict[str, str] = {}
    duplicate_of: dict[str, str] = {}
    converter = OpenCC("t2s.json")
    for checksum, paths in inventory.get("duplicate_groups", {}).items():
        sorted_paths = sorted(paths)
        for path in sorted_paths:
            group_for_path[path] = checksum
        # Identical files named for the same converted word are a recoverable
        # duplicate, not a claimed semantic cross-mapping.
        if len({normalize(converter.convert(Path(path).stem)) for path in sorted_paths}) == 1:
            for path in sorted_paths[1:]:
                duplicate_of[path] = sorted_paths[0]
    return group_for_path, duplicate_of


def reconcile_images(
    rows: list[dict[str, Any]],
    images: list[dict[str, Any]],
    inventory: dict[str, Any],
    production_by_source: dict[str, dict[str, Any]],
    production_ids_by_checksum: dict[str, list[str]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], set[str]]:
    """Return source-image records, row mapping metadata, and ambiguous rows."""
    converter = OpenCC("t2s.json")
    rows_by_word: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        rows_by_word[normalize(row["simplified"])].append(row)
    production_source_keys = set(production_by_source)
    group_for_path, duplicate_of = source_duplicate_metadata(inventory)

    candidates: dict[str, tuple[str, list[dict[str, Any]], str]] = {}
    target_to_images: dict[str, list[str]] = defaultdict(list)
    for image in images:
        relative_path = image["rel"]
        stem = normalize(Path(relative_path).stem)
        if image["sha256"] in production_ids_by_checksum:
            candidates[relative_path] = ("already-committed", [], stem)
            continue
        if relative_path == "拼音表.png":
            candidates[relative_path] = ("unsuitable-source-reference", [], stem)
            continue
        if relative_path in duplicate_of:
            candidates[relative_path] = ("duplicate-source", [], stem)
            continue
        raw_targets = rows_by_word.get(stem, [])
        converted_stem = normalize(converter.convert(stem))
        if raw_targets:
            evidence = "exact-normalized-basename"
            targets = raw_targets
        elif converted_stem != stem and converted_stem in rows_by_word:
            evidence = "unique-simplified-traditional-equivalence"
            targets = rows_by_word[converted_stem]
        else:
            evidence = "unmatched"
            targets = []
        targets = [target for target in targets if target["key"] not in production_source_keys]
        candidates[relative_path] = (evidence, targets, converted_stem)
        for target in targets:
            target_to_images[target["key"]].append(relative_path)

    source_records: list[dict[str, Any]] = []
    row_mapping: dict[str, dict[str, Any]] = {}
    ambiguous_rows: set[str] = set()
    for image in images:
        relative_path = image["rel"]
        evidence, targets, converted_stem = candidates[relative_path]
        checksum = image["sha256"]
        record: dict[str, Any] = {
            "id": f"teacher-image-{hashlib.sha256(relative_path.encode('utf-8')).hexdigest()[:16]}",
            "relativePath": relative_path,
            "sourceChecksumSha256": checksum,
            "width": image["width"],
            "height": image["height"],
            "format": "PNG",
            "evidence": evidence,
            "previewIds": [],
        }
        if relative_path in group_for_path:
            record["duplicateGroupChecksumSha256"] = group_for_path[relative_path]
        if evidence == "already-committed":
            record["state"] = "already-committed"
            preview_ids = production_ids_by_checksum[checksum]
            record["previewIds"] = preview_ids
            record["note"] = "Immutable production contract; existing production derivative is reused."
        elif evidence == "duplicate-source":
            record["state"] = "duplicate"
            record["duplicateOfRelativePath"] = duplicate_of[relative_path]
            record["note"] = "Byte-identical duplicate with the same converted basename; canonical source retained."
        elif evidence == "unsuitable-source-reference":
            record["state"] = "unsuitable"
            record["note"] = "734×1024 pinyin reference sheet, not a vocabulary illustration."
        elif not targets:
            record["state"] = "unmatched"
            record["note"] = "No exact basename, unique script-equivalence, or explicit mapping-note evidence."
        elif len(targets) != 1 or len(target_to_images[targets[0]["key"]]) != 1:
            record["state"] = "ambiguous"
            record["previewIds"] = []
            record["note"] = "Target row or candidate source image collision; no exact mapping claimed."
            for target in targets:
                ambiguous_rows.add(target["key"])
        else:
            target = targets[0]
            preview_id = target["previewId"]
            record["state"] = "mapped"
            record["previewIds"] = [preview_id]
            if evidence == "unique-simplified-traditional-equivalence":
                record["note"] = f"OpenCC t2s conversion: {Path(relative_path).stem} → {converted_stem}."
            row_mapping[target["key"]] = {
                "relativePath": relative_path,
                "sourceChecksumSha256": checksum,
                "evidence": evidence,
                "sourceWidth": image["width"],
                "sourceHeight": image["height"],
            }
        source_records.append(record)
    if len(source_records) != EXPECTED_IMAGE_COUNT:
        raise ValueError("Every readable teacher source image must have exactly one reconciliation record")
    return source_records, row_mapping, ambiguous_rows


def classify_missing_image(row: dict[str, Any], ambiguous_rows: set[str]) -> tuple[str, str]:
    word = row["simplified"]
    part_of_speech = row["partOfSpeech"]
    if row["key"] in ambiguous_rows or any(marker in word for marker in ("/", "、", ",", "，", ";", "；", "(", "（")):
        return "ambiguous", "The source expression has competing image candidates or multiple concepts."
    if word in SAFETY_SENSITIVE_TERMS:
        return "unsuitable", "The source term lacks a safely unambiguous single-scene illustration."
    if not re.search(r"[\u3400-\u9fff]", word):
        return "skipped", "The source value does not contain a usable Chinese concept."
    if word in TEXT_ONLY_TERMS or part_of_speech == "adverb":
        return "text-only", "Functional, deictic, or temporal source value is not safely rendered as one picture."
    if part_of_speech == "noun":
        if any(marker in word for marker in ABSTRACT_NOUN_MARKERS):
            return "unsuitable", "Abstract or culturally contextual noun is intentionally text-only."
        return "suitable", "Concrete noun candidate for an educational illustration."
    if part_of_speech == "verb":
        if word in FUNCTIONAL_VERBS or word.endswith(("起来", "下去", "过来", "过去")):
            return "text-only", "Functional, modal, or aspectual verb is not safely rendered as one picture."
        return "suitable", "Observable action candidate for an educational illustration."
    if part_of_speech == "adjective":
        if word in VISUAL_ADJECTIVES:
            return "suitable", "Directly observable attribute candidate for an educational illustration."
        return "unsuitable", "Non-visual or context-dependent adjective is intentionally text-only."
    return "unsuitable", "No safe illustration classification is available."


def prompt_for(row: dict[str, Any], generation_revision: int) -> str:
    subject = row["simplified"]
    revision_two_detail = (
        " Use an extra-thick, clearly hand-drawn outline; reduce small details; avoid glossy highlights."
        if generation_revision >= 2 else ""
    )
    revision_three_detail = (
        " Keep the full subject within the central 70% of the canvas; leave every border entirely #00ff00."
        if generation_revision >= 3 else ""
    )
    return "\n".join((
        "Use case: illustration-story",
        "Asset type: educational vocabulary preview illustration",
        f'Primary request: depict one clear, culturally neutral visual concept for the Chinese vocabulary item "{subject}".',
        "Style/medium: friendly hand-drawn digital cartoon, rounded dark charcoal-brown outline with mild variation, simple shapes, bright but not neon warm orange, grass green, lake blue, soft pink, and natural skin tones." + revision_two_detail + revision_three_detail,
        "Composition/framing: square 1:1, centered single concept, generous transparent-safe padding, at most one simple supporting context object.",
        "Constraints: simple educational composition, no written characters, no pinyin, no English, no letters, no logo, no watermark, no brand, no label, no quiz answer, no political symbol, no stereotype, no decorative clutter, no detailed background, no cast shadow.",
        "Output intent: isolated opaque subject on a perfectly flat #00ff00 chroma-key background for later alpha removal; use no #00ff00 in the subject.",
    ))


def image_input_for(preview_id: str, source_dir: Path | None) -> Path | None:
    if source_dir is None or not source_dir.exists():
        return None
    matches = [source_dir / f"{preview_id}{suffix}" for suffix in (".png", ".webp", ".jpg", ".jpeg")]
    found = [candidate for candidate in matches if candidate.is_file()]
    if len(found) > 1:
        raise ValueError(f"Multiple AI input images for {preview_id}")
    return found[0] if found else None


def verify_source_matches_mapping(
    source: Path,
    relative_path: str,
    expected_sha256: str,
    expected_width: int,
    expected_height: int,
) -> None:
    """Fail closed when the actual source file drifts from its reconciliation record."""
    actual_sha256 = sha256_file(source)
    if actual_sha256 != expected_sha256:
        raise ValueError(
            f"Teacher source '{relative_path}' checksum drift: "
            f"expected {expected_sha256}, got {actual_sha256}"
        )
    with Image.open(source) as opened:
        width, height = opened.size
    if (width, height) != (expected_width, expected_height):
        raise ValueError(
            f"Teacher source '{relative_path}' dimension drift: "
            f"expected {expected_width}x{expected_height}, got {width}x{height}"
        )


def export_derivative(source: Path, destination: Path, *, require_transparent_corners: bool = False) -> dict[str, Any]:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as opened:
        image = ImageOps.exif_transpose(opened)
        if image.mode not in ("RGB", "RGBA"):
            image = image.convert("RGBA" if "A" in image.getbands() else "RGB")
        if require_transparent_corners:
            validate_transparent_corners(image, source)
        image.thumbnail((512, 512), Image.Resampling.LANCZOS)
        image.save(destination, format="WEBP", lossless=True, method=6, exif=b"", icc_profile=None)
    with Image.open(destination) as output:
        width, height = output.size
    return {"assetChecksumSha256": sha256_file(destination), "width": width, "height": height}


def validate_transparent_corners(image: Image.Image, source: Path) -> None:
    if image.mode != "RGBA":
        raise ValueError(f"AI input '{source.name}' has no alpha channel; remove its chroma-key background first")
    alpha = image.getchannel("A")
    corners = ((0, 0), (image.width - 1, 0), (0, image.height - 1), (image.width - 1, image.height - 1))
    if any(alpha.getpixel(corner) > 32 for corner in corners):
        raise ValueError(f"AI input '{source.name}' does not have transparent corners")


def validate_ai_input(source: Path) -> None:
    with Image.open(source) as image:
        validate_transparent_corners(image, source)


def existing_derivative_metadata(destination: Path) -> dict[str, Any] | None:
    """Read a prior deterministic derivative without rewriting it."""
    try:
        with Image.open(destination) as output:
            if output.format != "WEBP":
                return None
            width, height = output.size
    except (OSError, ValueError):
        return None
    return {"assetChecksumSha256": sha256_file(destination), "width": width, "height": height}


def resolve_teacher_derivative(
    source: Path,
    relative_path: str,
    expected_sha256: str,
    expected_width: int,
    expected_height: int,
    destination: Path,
    *,
    rebuild: bool,
    label: str,
) -> dict[str, Any]:
    """Verify the source then reuse or regenerate the derivative.

    Shared by the fresh-export and reuse paths so both carry the same source
    checksum/dimension gate and the same derivative byte-integrity gate.
    """
    verify_source_matches_mapping(
        source, relative_path, expected_sha256, expected_width, expected_height
    )
    if rebuild:
        return export_derivative(source, destination)
    return resolve_derivative(
        source, destination, require_transparent_corners=False, label=label
    )


def resolve_derivative(
    source: Path,
    destination: Path,
    *,
    require_transparent_corners: bool,
    label: str,
) -> dict[str, Any]:
    """Reuse an existing derivative only when it matches a deterministic re-export.

    The existing WebP is accepted only when re-exporting the current source
    produces byte-identical output. A changed source, corrupt/unreadable file,
    or a derivative copied from another preview item fails the byte comparison
    and the derivative is regenerated deterministically from the verified
    source. Fresh export and reuse therefore share the same integrity contract.
    """
    existing = existing_derivative_metadata(destination)
    with tempfile.TemporaryDirectory() as tmp:
        expected = Path(tmp) / destination.name
        rendered = export_derivative(
            source, expected, require_transparent_corners=require_transparent_corners
        )
        expected_checksum = rendered["assetChecksumSha256"]
        if existing is None or existing["assetChecksumSha256"] != expected_checksum:
            # The existing file is absent, unreadable, a non-WebP, or drifted;
            # overwrite it deterministically from the verified source.
            if existing is not None:
                print(f"  REUSE-REJECT  {label}: existing derivative did not match a deterministic re-export")
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(expected, destination)
            return rendered
    return existing


def load_accepted_ai_records(output_dir: Path) -> dict[str, dict[str, Any]]:
    """Load accepted AI-generated records from the committed preview corpus.

    Returns preview_id -> image record for every row whose state is
    ai-generated. These are the authoritative expected assets for the
    --reuse-accepted-ai-assets mode.
    """
    corpus = json.loads((output_dir / "preview-corpus.json").read_text(encoding="utf-8"))
    accepted: dict[str, dict[str, Any]] = {}
    for row in corpus["rows"]:
        image = row.get("image", {})
        if image.get("state") == "ai-generated":
            if not image.get("assetPath") or not image.get("assetChecksumSha256"):
                raise ValueError(f"Accepted AI row '{row['id']}' is missing assetPath or checksum")
            accepted[row["id"]] = image
    if not accepted:
        raise ValueError("No accepted AI-generated records found in the committed preview corpus")
    return accepted


def verify_accepted_ai_asset(
    preview_id: str,
    record: dict[str, Any],
    ai_asset_dir: Path,
    seen_paths: dict[str, str],
    *,
    prompt_digest: str,
    generation_revision: int,
    reference_set_ids: tuple[str, ...],
) -> dict[str, Any]:
    """Verify one accepted AI asset fails closed, then return its metadata.

    Checks expected preview ID, file existence, WebP readability and
    dimensions, SHA-256 against the accepted record, transparent-corner
    requirement, a one-to-one ID/path mapping, and that the committed
    generation provenance (promptDigest, generationRevision, referenceSetIds)
    still matches the values currently derived from the frozen prompt contract.
    A drift means the on-disk asset was produced by a different prompt than the
    current build would claim, so reuse fails closed. Accepted assets are never
    rewritten.
    """
    asset_path = record["assetPath"]
    expected_filename = f"{preview_id}.webp"
    if Path(asset_path).name != expected_filename:
        raise ValueError(f"Accepted AI preview '{preview_id}' assetPath '{asset_path}' does not match its preview ID")
    if asset_path in seen_paths:
        raise ValueError(f"Accepted AI assetPath '{asset_path}' is shared by previews '{seen_paths[asset_path]}' and '{preview_id}'")
    # Provenance must agree with the currently frozen prompt contract.
    if record.get("promptDigest") != prompt_digest:
        raise ValueError(
            f"Accepted AI preview '{preview_id}' promptDigest drift: "
            f"expected {prompt_digest}, got {record.get('promptDigest')}"
        )
    if record.get("generationRevision") != generation_revision:
        raise ValueError(
            f"Accepted AI preview '{preview_id}' generationRevision drift: "
            f"expected {generation_revision}, got {record.get('generationRevision')}"
        )
    if tuple(record.get("referenceSetIds") or ()) != reference_set_ids:
        raise ValueError(
            f"Accepted AI preview '{preview_id}' referenceSetIds drift: "
            f"expected {list(reference_set_ids)}, got {record.get('referenceSetIds')}"
        )
    destination = ai_asset_dir / expected_filename
    if not destination.is_file():
        raise ValueError(f"Accepted AI asset missing: {destination}")
    try:
        with Image.open(destination) as opened:
            if opened.format != "WEBP":
                raise ValueError(f"Accepted AI asset '{destination.name}' is not WebP")
            width, height = opened.size
    except (OSError, ValueError) as exc:
        raise ValueError(f"Accepted AI asset unreadable: {destination} ({exc})") from exc
    expected_width, expected_height = record["width"], record["height"]
    if (width, height) != (expected_width, expected_height):
        raise ValueError(
            f"Accepted AI asset '{destination.name}' dimension drift: "
            f"expected {expected_width}x{expected_height}, got {width}x{height}"
        )
    actual_sha256 = sha256_file(destination)
    if actual_sha256 != record["assetChecksumSha256"]:
        raise ValueError(
            f"Accepted AI asset '{destination.name}' checksum drift: "
            f"expected {record['assetChecksumSha256']}, got {actual_sha256}"
        )
    validate_transparent_corners(Image.open(destination), destination)
    seen_paths[asset_path] = preview_id
    return {"assetChecksumSha256": actual_sha256, "width": width, "height": height}


def missing_fields(row: dict[str, Any]) -> list[str]:
    fields = [field for field in ("pinyin", "japanese", "difficulty") if not row[field]]
    if not row.get("traditional"):
        fields.append("traditional")
    return fields


def build(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = args.workbook.resolve()
    source_dir = args.source_dir.resolve()
    output_dir = args.output_dir.resolve()
    teacher_asset_dir = args.teacher_asset_dir.resolve()
    ai_asset_dir = args.ai_asset_dir.resolve()
    ai_source_dir = args.ai_source_dir.resolve() if args.ai_source_dir else None
    accepted_ai_records = None
    seen_ai_paths: dict[str, str] = {}
    if args.reuse_accepted_ai_assets:
        if ai_source_dir is not None:
            raise ValueError("--reuse-accepted-ai-assets and --ai-source-dir are mutually exclusive")
        if args.rebuild_ai_assets:
            raise ValueError("--reuse-accepted-ai-assets and --rebuild-ai-assets are mutually exclusive")
        accepted_ai_records = load_accepted_ai_records(output_dir)
    if not source_dir.is_dir():
        raise ValueError(f"Teacher image source directory does not exist: {source_dir}")
    images, inventory = load_inventory(args.inventory.resolve())
    rows = load_workbook_rows(workbook_path)
    production_by_source, production_ids_by_checksum = production_records()
    for row in rows:
        row["previewId"] = production_by_source.get(row["key"], {}).get("id") or stable_preview_id(
            row["sourceSheet"], row["sourceRow"], row["simplified"]
        )
    reconciliation, row_mapping, ambiguous_rows = reconcile_images(
        rows, images, inventory, production_by_source, production_ids_by_checksum
    )

    queue: list[dict[str, Any]] = []
    corpus_rows: list[dict[str, Any]] = []
    for row in rows:
        production = production_by_source.get(row["key"])
        preview_row = {
            "id": row["previewId"],
            "simplified": row["simplified"],
            "pinyin": row["pinyin"] or None,
            "japanese": row["japanese"] or None,
            "traditional": None,
            "difficulty": row["difficulty"] or None,
            "partOfSpeech": row["partOfSpeech"],
            "sourceSheet": row["sourceSheet"],
            "sourceRow": row["sourceRow"],
            "reviewStatus": "draft",
        }
        if production:
            preview_row.update({
                "productionVocabularyId": production["id"],
                "simplified": production["simplified"],
                "pinyin": production["pinyin"],
                "japanese": production["japanese"],
                "traditional": production.get("traditional"),
                "difficulty": production["curriculum"]["sourceDifficultyLabel"],
            })
            illustration = next((
                item for item in json.loads(
                    (REPO_ROOT / "data/illustrations/teacher-core-v1/teacher-vocabulary-batch-01.json").read_text(encoding="utf-8")
                )["illustrations"] if item["vocabularyId"] == production["id"]
            ), None)
            if illustration:
                preview_row["image"] = {
                    "state": "teacher-mapped",
                    "provenance": "teacher-provided",
                    "reviewStatus": "draft",
                    "assetPath": illustration["assetPath"],
                    "sourceChecksumSha256": illustration["sourceChecksumSha256"],
                    "width": illustration["width"],
                    "height": illustration["height"],
                    "reconciliationEvidence": "existing-immutable-production-contract",
                }
            else:
                preview_row["image"] = {
                    "state": "text-only", "provenance": None, "reviewStatus": "not-applicable",
                    "note": "Immutable production contract preserves this text-only row.",
                }
        elif row["key"] in row_mapping:
            mapping = row_mapping[row["key"]]
            destination = teacher_asset_dir / f"{row['previewId']}.webp"
            # Verify the actual source against its reconciliation record and
            # reuse the existing derivative only when it matches a deterministic
            # re-export, in both the fresh-export and the reuse path.
            exported = resolve_teacher_derivative(
                source_dir / mapping["relativePath"],
                mapping["relativePath"],
                mapping["sourceChecksumSha256"],
                mapping["sourceWidth"],
                mapping["sourceHeight"],
                destination,
                rebuild=args.rebuild_teacher_assets,
                label=row["previewId"],
            )
            preview_row["image"] = {
                "state": "teacher-mapped",
                "provenance": "teacher-provided",
                "reviewStatus": "draft",
                "assetPath": f"/assets/vocabulary/teacher-preview/teacher/{row['previewId']}.webp",
                "sourceImageRelativePath": mapping["relativePath"],
                "sourceChecksumSha256": mapping["sourceChecksumSha256"],
                "reconciliationEvidence": mapping["evidence"],
                **exported,
            }
        else:
            state, reason = classify_missing_image(row, ambiguous_rows)
            if state == "suitable":
                generation_revision = (
                    3 if row["previewId"] in REGENERATED_PREVIEW_IDS
                    else 1 if row["previewId"] in CHECKPOINT_PREVIEW_IDS else 2
                )
                prompt = prompt_for(row, generation_revision)
                digest = hashlib.sha256(prompt.encode("utf-8")).hexdigest()
                if accepted_ai_records is not None:
                    # Reuse mode: verify the existing accepted AI asset against
                    # the committed record and preserve it without rewriting.
                    record = accepted_ai_records.get(row["previewId"])
                    if record is None:
                        raise ValueError(
                            f"Reuse mode: preview '{row['previewId']}' is expected to have an accepted AI asset "
                            "but none is recorded in the committed preview corpus"
                        )
                    exported = verify_accepted_ai_asset(
                        row["previewId"], record, ai_asset_dir, seen_ai_paths,
                        prompt_digest=digest,
                        generation_revision=generation_revision,
                        reference_set_ids=STYLE_REFERENCE_SET_IDS,
                    )
                    preview_row["image"] = {
                        "state": "ai-generated", "provenance": "ai-generated", "reviewStatus": "draft",
                        "assetPath": f"/assets/vocabulary/teacher-preview/ai/{row['previewId']}.webp",
                        "promptDigest": digest,
                        "generationRevision": generation_revision,
                        "referenceSetIds": STYLE_REFERENCE_SET_IDS,
                        "reviewOutcome": "automated-alpha-check; draft human review pending",
                        "reconciliationEvidence": "frozen-generation-queue", **exported,
                    }
                    queue_status = "generated"
                else:
                    generated_input = image_input_for(row["previewId"], ai_source_dir)
                    if generated_input:
                        destination = ai_asset_dir / f"{row['previewId']}.webp"
                        validate_ai_input(generated_input)
                        exported = None if args.rebuild_ai_assets or not destination.is_file() else existing_derivative_metadata(destination)
                        if exported is None:
                            exported = export_derivative(generated_input, destination, require_transparent_corners=True)
                        preview_row["image"] = {
                            "state": "ai-generated", "provenance": "ai-generated", "reviewStatus": "draft",
                            "assetPath": f"/assets/vocabulary/teacher-preview/ai/{row['previewId']}.webp",
                            "promptDigest": digest,
                            "generationRevision": generation_revision,
                            "referenceSetIds": STYLE_REFERENCE_SET_IDS,
                            "reviewOutcome": "automated-alpha-check; draft human review pending",
                            "reconciliationEvidence": "frozen-generation-queue", **exported,
                        }
                        queue_status = "generated"
                    else:
                        preview_row["image"] = {
                            "state": "ai-pending", "provenance": None, "reviewStatus": "not-applicable",
                            "promptDigest": digest, "note": reason,
                        }
                        queue_status = "pending"
                queue.append({
                    "previewId": row["previewId"], "sourceSheet": row["sourceSheet"], "sourceRow": row["sourceRow"],
                    "simplified": row["simplified"], "partOfSpeech": row["partOfSpeech"], "prompt": prompt,
                    "promptDigest": digest, "generationRevision": generation_revision, "status": queue_status,
                    **({"regenerationReason": REGENERATION_REASONS[row["previewId"]]}
                       if row["previewId"] in REGENERATED_PREVIEW_IDS else {}),
                })
            else:
                preview_row["image"] = {
                    "state": state, "provenance": None, "reviewStatus": "not-applicable", "note": reason,
                }
        preview_row["missingFields"] = missing_fields(preview_row)
        # JSON has no undefined; explicit null values are removed after missing-field capture.
        for key in ("pinyin", "japanese", "traditional", "difficulty"):
            if preview_row[key] is None:
                del preview_row[key]
        corpus_rows.append(preview_row)

    by_image_state = Counter(row["image"]["state"] for row in corpus_rows)
    totals = {
        "usableRows": len(corpus_rows),
        "missingFields": {field: sum(field in row["missingFields"] for row in corpus_rows)
                          for field in ("pinyin", "japanese", "traditional", "difficulty")},
        "bySourceSheet": dict(Counter(row["sourceSheet"] for row in corpus_rows)),
        "byPartOfSpeech": dict(Counter(row["partOfSpeech"] for row in corpus_rows)),
        "byImageState": {state: by_image_state.get(state, 0) for state in (
            "teacher-mapped", "ai-generated", "ai-pending", "text-only", "ambiguous", "unsuitable", "skipped"
        )},
    }
    corpus = {
        "schemaVersion": 1,
        "workbook": {"basename": workbook_path.name, "sha256": EXPECTED_WORKBOOK_SHA256, "candidateRows": 1865},
        "teacherImagePackage": {"readableImages": EXPECTED_IMAGE_COUNT, "pathSensitiveFingerprintSha256": EXPECTED_IMAGE_FINGERPRINT},
        "totals": totals,
        "rows": corpus_rows,
    }
    reconciliation_payload = {
        "schemaVersion": 1,
        "input": {"readableImages": EXPECTED_IMAGE_COUNT, "pathSensitiveFingerprintSha256": EXPECTED_IMAGE_FINGERPRINT},
        "stateTotals": dict(Counter(record["state"] for record in reconciliation)),
        "images": reconciliation,
    }
    queue_payload = {
        "schemaVersion": 1,
        "styleAudit": "docs/teacher-vocabulary-preview/style-audit.md",
        "status": "complete" if not any(item["status"] == "pending" for item in queue) else "generation-required",
        "totals": {
            "suitable": len(queue),
            "generated": sum(item["status"] == "generated" for item in queue),
            "pending": sum(item["status"] == "pending" for item in queue),
            "rejected": sum(item["status"] == "rejected" for item in queue),
            "regenerated": sum(1 for item in queue if item.get("regenerationReason")),
        },
        "items": queue,
    }
    archives = [
        {"relativePath": record["rel"], "sizeBytes": record["size"], "sha256": record["sha256"]}
        for record in inventory.get("records", []) if record.get("ext") == ".rar"
    ]
    anomaly = next(
        record for record in images if record.get("width") == 734 and record.get("height") == 1024
    )
    input_audit = {
        "schemaVersion": 1,
        "workbook": {"basename": workbook_path.name, "sha256": EXPECTED_WORKBOOK_SHA256, "candidateRows": 1865},
        "teacherImagePackage": {
            "readableImages": EXPECTED_IMAGE_COUNT,
            "pathSensitiveFingerprintSha256": EXPECTED_IMAGE_FINGERPRINT,
            "duplicateGroups": inventory["duplicate_groups"],
            "dimensionCounts": {
                "500x500": sum(record.get("width") == 500 and record.get("height") == 500 for record in images),
                "500x501": sum(record.get("width") == 500 and record.get("height") == 501 for record in images),
                "501x500": sum(record.get("width") == 501 and record.get("height") == 500 for record in images),
                "501x501": sum(record.get("width") == 501 and record.get("height") == 501 for record in images),
                "734x1024": 1,
            },
            "dimensionAnomaly": {
                "relativePath": anomaly["rel"], "sha256": anomaly["sha256"],
                "width": anomaly["width"], "height": anomaly["height"],
                "classification": "unsuitable source reference image",
            },
            "archives": archives,
            "archiveHandling": "Inspected by inventory only. No archive was extracted into a tracked path.",
        },
    }
    atomic_json(output_dir / "preview-corpus.json", corpus)
    atomic_json(output_dir / "teacher-image-reconciliation.json", reconciliation_payload)
    atomic_json(output_dir / "generation-queue.json", queue_payload)
    atomic_json(output_dir / "input-inventory-audit.json", input_audit)
    return {"corpus": corpus, "reconciliation": reconciliation_payload, "queue": queue_payload, "inputAudit": input_audit}


def make_transparent_webp(path: Path) -> bytes:
    """Write a deterministic RGBA 512x512 WebP with transparent corners."""
    from PIL import ImageDraw
    image = Image.new("RGBA", (512, 512), (0, 0, 0, 0))
    ImageDraw.Draw(image).rectangle((32, 32, 479, 479), fill=(120, 160, 200, 255))
    image.save(path, format="WEBP", lossless=True, method=6, exif=b"", icc_profile=None)
    return path.read_bytes()


def run_self_tests() -> int:
    """Read-only drift tests for source and derivative integrity (Issues #185/#192)."""
    import tempfile

    failures = 0

    def expect_raises(description: str, expected: str, fn: Any) -> None:
        nonlocal failures
        try:
            fn()
            print(f"  FAIL  {description}: expected ValueError containing {expected!r}, none raised")
            failures += 1
        except ValueError as exc:
            if expected in str(exc):
                print(f"  PASS  {description}")
            else:
                print(f"  FAIL  {description}: unexpected message {exc}")
                failures += 1

    with tempfile.TemporaryDirectory() as tmp:
        base = Path(tmp) / "source.png"
        with Image.new("RGB", (50, 60), (255, 0, 0)) as image:
            image.save(base)
        original_sha256 = sha256_file(base)

        def pass_case() -> None:
            verify_source_matches_mapping(base, "source.png", original_sha256, 50, 60)

        try:
            pass_case()
            print("  PASS  verify_source_matches_mapping accepts matching checksum and dimensions")
        except ValueError as exc:
            print(f"  FAIL  matching source rejected: {exc}")
            failures += 1

        def checksum_drift_case() -> None:
            with Image.new("RGB", (50, 60), (0, 255, 0)) as image:
                image.save(base)
            try:
                verify_source_matches_mapping(base, "source.png", original_sha256, 50, 60)
            finally:
                with Image.new("RGB", (50, 60), (255, 0, 0)) as image:
                    image.save(base)

        expect_raises("checksum drift fails closed", "checksum drift", checksum_drift_case)

        def dimension_drift_case() -> None:
            # Pass the file's real checksum but a wrong dimension so only the
            # dimension gate is triggered (a changed size always changes the
            # checksum too, so this exercises the independent dimension check).
            verify_source_matches_mapping(base, "source.png", sha256_file(base), 99, 99)

        expect_raises("dimension drift fails closed", "dimension drift", dimension_drift_case)

        try:
            pass_case()
            print("  PASS  source is restored and still matches after drift cases")
        except ValueError as exc:
            print(f"  FAIL  source restore verification failed: {exc}")
            failures += 1

        # ── Reuse-path derivative integrity (Issue #192) ──
        # resolve_teacher_derivative must reuse a matching derivative as-is and
        # fail closed on a drifted source, corrupt or mismatched existing WebP,
        # and a derivative copied from another preview item.
        derivative = Path(tmp) / "teacher-preview-0123456789abcdef.webp"
        fresh = resolve_teacher_derivative(
            base, "source.png", original_sha256, 50, 60, derivative,
            rebuild=False, label="reuse-test",
        )
        fresh_checksum = fresh["assetChecksumSha256"]

        # 1. Matching source + matching existing derivative is reused as-is.
        reused = resolve_teacher_derivative(
            base, "source.png", original_sha256, 50, 60, derivative,
            rebuild=False, label="reuse-test",
        )
        if reused == fresh and derivative.exists():
            print("  PASS  matching derivative is reused successfully")
        else:
            print("  FAIL  matching derivative was not reused as-is")
            failures += 1

        # 2. Source checksum drift fails in the reuse path.
        with Image.new("RGB", (50, 60), (0, 255, 0)) as image:
            image.save(base)
        try:
            resolve_teacher_derivative(
                base, "source.png", original_sha256, 50, 60, derivative,
                rebuild=False, label="reuse-test",
            )
            print("  FAIL  source checksum drift did not raise")
            failures += 1
        except ValueError as exc:
            if "checksum drift" in str(exc):
                print("  PASS  source checksum drift fails in the reuse path")
            else:
                print(f"  FAIL  unexpected checksum-drift message: {exc}")
                failures += 1
        finally:
            with Image.new("RGB", (50, 60), (255, 0, 0)) as image:
                image.save(base)

        # 3. Source dimension drift fails in the reuse path.
        try:
            resolve_teacher_derivative(
                base, "source.png", original_sha256, 99, 99, derivative,
                rebuild=False, label="reuse-test",
            )
            print("  FAIL  source dimension drift did not raise")
            failures += 1
        except ValueError as exc:
            if "dimension drift" in str(exc):
                print("  PASS  source dimension drift fails in the reuse path")
            else:
                print(f"  FAIL  unexpected dimension-drift message: {exc}")
                failures += 1

        # 4. Corrupt existing WebP is deterministically regenerated.
        corrupt = Path(tmp) / "corrupt.webp"
        corrupt.write_bytes(b"not a webp")
        regenerated = resolve_teacher_derivative(
            base, "source.png", original_sha256, 50, 60, corrupt,
            rebuild=False, label="corrupt-test",
        )
        if regenerated["assetChecksumSha256"] == fresh_checksum and Image.open(corrupt).format == "WEBP":
            print("  PASS  corrupt existing WebP is deterministically regenerated")
        else:
            print("  FAIL  corrupt WebP regeneration failed")
            failures += 1

        # 5. A valid WebP copied from another preview item is rejected and
        #    overwritten with the expected deterministic output.
        foreign = Path(tmp) / "foreign.webp"
        with Image.new("RGB", (80, 80), (0, 0, 255)) as image:
            image.save(foreign, format="WEBP", lossless=True)
        foreign_checksum = sha256_file(foreign)
        fixed = resolve_teacher_derivative(
            base, "source.png", original_sha256, 50, 60, foreign,
            rebuild=False, label="cross-copied-test",
        )
        if fixed["assetChecksumSha256"] == fresh_checksum and sha256_file(foreign) != foreign_checksum:
            print("  PASS  WebP copied from another item is rejected and replaced")
        else:
            print("  FAIL  cross-copied WebP was not rejected")
            failures += 1

        # 6. Altered derivative bytes are detected and deterministically replaced.
        altered_bytes = bytearray(derivative.read_bytes())
        altered_bytes[:8] = b"\x00" * 8
        derivative.write_bytes(bytes(altered_bytes))
        restored = resolve_teacher_derivative(
            base, "source.png", original_sha256, 50, 60, derivative,
            rebuild=False, label="altered-test",
        )
        if restored["assetChecksumSha256"] == fresh_checksum:
            print("  PASS  altered derivative bytes are detected and regenerated")
        else:
            print("  FAIL  altered derivative bytes were not detected")
            failures += 1

        # 7. Recorded metadata agrees with the actual output bytes.
        with Image.open(derivative) as output:
            actual_width, actual_height = output.size
        if (
            sha256_file(derivative) == restored["assetChecksumSha256"]
            and actual_width == restored["width"]
            and actual_height == restored["height"]
        ):
            print("  PASS  recorded metadata matches the actual output bytes")
        else:
            print("  FAIL  recorded metadata does not match the actual output bytes")
            failures += 1

        # ── Accepted-AI reuse verification (Issue #193 follow-up) ──
        # verify_accepted_ai_asset must accept a matching accepted asset as-is
        # and fail closed on missing/corrupt/drifted/cross-copied assets or
        # drifted generation provenance.
        ai_dir = Path(tmp) / "ai"
        ai_dir.mkdir()
        seen_ai: dict[str, str] = {}
        ai_pid = "teacher-preview-accepted-ai"
        ai_asset = ai_dir / f"{ai_pid}.webp"
        ai_asset_bytes = make_transparent_webp(ai_asset)
        ai_prov = {
            "promptDigest": "a" * 64,
            "generationRevision": 2,
            "referenceSetIds": ["one.png", "two.png", "three.png"],
        }
        ai_record = {
            "assetPath": f"/assets/vocabulary/teacher-preview/ai/{ai_pid}.webp",
            "assetChecksumSha256": sha256_file(ai_asset),
            "width": 512,
            "height": 512,
            **ai_prov,
        }

        def reuse_accepted() -> dict[str, Any]:
            return verify_accepted_ai_asset(
                ai_pid, ai_record, ai_dir, seen_ai,
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            )

        # a. Matching accepted asset is reused as-is (bytes unchanged).
        try:
            reused_meta = reuse_accepted()
            if reused_meta["assetChecksumSha256"] == ai_record["assetChecksumSha256"] and ai_asset.read_bytes() == ai_asset_bytes:
                print("  PASS  accepted AI asset is reused as-is")
            else:
                print("  FAIL  accepted AI asset was not reused as-is")
                failures += 1
        except ValueError as exc:
            print(f"  FAIL  accepted AI asset reuse raised: {exc}")
            failures += 1

        # b. Missing asset fails closed.
        missing_pid = "teacher-preview-missing-ai"
        missing_record = {
            "assetPath": f"/assets/vocabulary/teacher-preview/ai/{missing_pid}.webp",
            "assetChecksumSha256": "0" * 64, "width": 512, "height": 512, **ai_prov,
        }
        expect_raises(
            "missing accepted AI asset fails closed",
            "missing",
            lambda: verify_accepted_ai_asset(
                missing_pid, missing_record, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # c. Checksum drift fails closed (a valid but different WebP at the path).
        drifted = ai_dir / f"{ai_pid}.webp"
        from PIL import ImageDraw as _ImageDraw
        drift_image = Image.new("RGBA", (512, 512), (0, 0, 0, 0))
        _ImageDraw.Draw(drift_image).rectangle((32, 32, 479, 479), fill=(10, 20, 30, 255))
        drift_image.save(drifted, format="WEBP", lossless=True, method=6, exif=b"", icc_profile=None)
        expect_raises(
            "accepted AI checksum drift fails closed",
            "checksum drift",
            lambda: verify_accepted_ai_asset(
                ai_pid, ai_record, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )
        ai_asset.write_bytes(ai_asset_bytes)

        # d. Dimension drift fails closed.
        expect_raises(
            "accepted AI dimension drift fails closed",
            "dimension drift",
            lambda: verify_accepted_ai_asset(
                ai_pid, {**ai_record, "width": 999}, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # e. Non-WebP format at the expected path fails closed.
        png_path = ai_dir / "teacher-preview-png.webp"
        with Image.new("RGB", (512, 512), (0, 0, 255)) as image:
            image.save(png_path, format="PNG")
        png_record = {
            "assetPath": "/assets/vocabulary/teacher-preview/ai/teacher-preview-png.webp",
            "assetChecksumSha256": sha256_file(png_path), "width": 512, "height": 512, **ai_prov,
        }
        expect_raises(
            "non-WebP accepted AI asset fails closed",
            "not WebP",
            lambda: verify_accepted_ai_asset(
                "teacher-preview-png", png_record, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # f. Transparent-corner regression fails closed (RGB has no alpha channel).
        opaque = ai_dir / "teacher-preview-opaque.webp"
        with Image.new("RGB", (512, 512), (0, 0, 255)) as image:
            image.save(opaque, format="WEBP", lossless=True)
        opaque_record = {
            "assetPath": "/assets/vocabulary/teacher-preview/ai/teacher-preview-opaque.webp",
            "assetChecksumSha256": sha256_file(opaque), "width": 512, "height": 512, **ai_prov,
        }
        expect_raises(
            "accepted AI transparent-corner regression fails closed",
            "alpha",
            lambda: verify_accepted_ai_asset(
                "teacher-preview-opaque", opaque_record, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # g. Cross-copied asset fails closed: a record pointing at another
        #    preview's asset path is rejected by the one-to-one ID/path mapping.
        cross_asset = ai_dir / "teacher-preview-cross.webp"
        cross_asset.write_bytes(ai_asset_bytes)
        expect_raises(
            "accepted AI cross-copied asset fails closed",
            "does not match its preview ID",
            lambda: verify_accepted_ai_asset(
                "teacher-preview-cross", {**ai_record}, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # h. Prompt digest drift fails closed.
        expect_raises(
            "accepted AI promptDigest drift fails closed",
            "promptDigest drift",
            lambda: verify_accepted_ai_asset(
                ai_pid, {**ai_record, "promptDigest": "b" * 64}, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # i. Generation revision drift fails closed.
        expect_raises(
            "accepted AI generationRevision drift fails closed",
            "generationRevision drift",
            lambda: verify_accepted_ai_asset(
                ai_pid, {**ai_record, "generationRevision": 3}, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

        # j. Reference-set drift fails closed.
        expect_raises(
            "accepted AI referenceSetIds drift fails closed",
            "referenceSetIds drift",
            lambda: verify_accepted_ai_asset(
                ai_pid, {**ai_record, "referenceSetIds": ["four.png"]}, ai_dir, {},
                prompt_digest=ai_prov["promptDigest"],
                generation_revision=ai_prov["generationRevision"],
                reference_set_ids=tuple(ai_prov["referenceSetIds"]),
            ),
        )

    if failures:
        print(f"\n{failures} integrity test(s) FAILED")
    else:
        print("\nAll source/derivative/AI-reuse integrity tests PASSED")
    return failures


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the complete #185 teacher vocabulary preview")
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--source-dir", type=Path)
    parser.add_argument("--inventory", type=Path)
    parser.add_argument("--output-dir", type=Path, default=REPO_ROOT / "data/teacher-vocabulary-preview")
    parser.add_argument("--teacher-asset-dir", type=Path, default=REPO_ROOT / "public/assets/vocabulary/teacher-preview/teacher")
    parser.add_argument("--ai-asset-dir", type=Path, default=REPO_ROOT / "public/assets/vocabulary/teacher-preview/ai")
    parser.add_argument("--ai-source-dir", type=Path)
    parser.add_argument(
        "--rebuild-teacher-assets",
        action="store_true",
        help="Regenerate every deterministic teacher preview derivative instead of reusing valid WebP outputs.",
    )
    parser.add_argument(
        "--rebuild-ai-assets",
        action="store_true",
        help="Regenerate every accepted AI preview derivative instead of reusing valid WebP outputs.",
    )
    parser.add_argument(
        "--reuse-accepted-ai-assets",
        action="store_true",
        help=(
            "Verify the existing accepted AI assets against the committed preview corpus and "
            "preserve them as ai-generated without rewriting. Mutually exclusive with "
            "--ai-source-dir and --rebuild-ai-assets."
        ),
    )
    parser.add_argument("--test", action="store_true", help="Run read-only source-verification drift tests")
    parser.add_argument("--build", action="store_true")
    args = parser.parse_args()
    if args.test:
        return run_self_tests()
    if not args.build:
        parser.error("--build is required; this explicit mode prevents accidental writes")
    if not (args.workbook and args.source_dir and args.inventory):
        parser.error("--workbook, --source-dir and --inventory are required in build mode")
    result = build(args)
    print(json.dumps({
        "usableRows": result["corpus"]["totals"]["usableRows"],
        "imageStates": result["corpus"]["totals"]["byImageState"],
        "reconciliation": result["reconciliation"]["stateTotals"],
        "generation": result["queue"]["totals"],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
