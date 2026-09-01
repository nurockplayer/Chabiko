"""Teacher-specific authoring sidecar contract for Issue #478.

This module reads the workbook's logical ``造词/造句`` value before the
learner-manifest normalization path.  Its output is authoring evidence only;
runtime loaders and the learner Unicode manifest must not import it.
"""

from __future__ import annotations

import copy
import hashlib
import json
import os
import re
import tempfile
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CONTRACT_ID = "teacher-phrase-authoring-v1"
PHRASE_ID_DOMAIN = "teacher-phrase-v1"
SOURCE_REVISION_DOMAIN = "teacher-phrase-source-v1"
SOURCE_COLUMN = "造词/造句"
RIGHTS_REF = "https://github.com/nurockplayer/Chabiko/issues/340#issuecomment-5279951072"


class ContractError(ValueError):
    """Raised when authoring evidence does not match its frozen source."""


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def serialize_sidecar(sidecar: dict[str, Any]) -> bytes:
    return (json.dumps(sidecar, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def atomic_write_sidecar(path: Path, sidecar: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(serialize_sidecar(sidecar))
            handle.flush()
            os.fsync(handle.fileno())
        temporary_path.replace(path)
    except BaseException:
        temporary_path.unlink(missing_ok=True)
        raise


def normalize_example(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", value).strip())


def manifest_semantic_sha256(manifest: dict[str, Any]) -> str:
    semantic = {
        "schemaVersion": manifest.get("schemaVersion"),
        "rows": [
            {
                "learnerId": row.get("learnerId"),
                "sourceSheet": row.get("sourceSheet"),
                "sourceRow": row.get("sourceRow"),
                **({"example": row["example"]} if "example" in row else {}),
            }
            for row in manifest.get("rows", [])
        ],
    }
    encoded = json.dumps(
        semantic,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return sha256_bytes(encoded)


def validate_manifest_workbook_digest(manifest: dict[str, Any], workbook_sha256: str) -> None:
    source = manifest.get("source")
    expected = source.get("workbookSha256") if isinstance(source, dict) else None
    if not isinstance(expected, str) or not re.fullmatch(r"[0-9a-f]{64}", expected):
        raise ContractError("learner manifest source is missing a valid workbook digest")
    if expected != workbook_sha256:
        raise ContractError(
            f"manifest workbook digest mismatch: expected {expected}, got {workbook_sha256}"
        )


def _domain_digest(domain: str, *values: object) -> str:
    payload = json.dumps(
        [domain, *values],
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return sha256_bytes(payload)


def phrase_id(learner_id: str, simplified: str, duplicate_discriminator: str = "") -> str:
    semantic_unit = unicodedata.normalize("NFC", simplified.strip())
    return f"teacher-phrase-v1-{_domain_digest(PHRASE_ID_DOMAIN, learner_id, semantic_unit, duplicate_discriminator)}"


def source_revision(
    learner_id: str,
    sheet: str,
    row: int,
    raw_cell_sha256: str,
) -> str:
    return f"teacher-phrase-source-v1-{_domain_digest(SOURCE_REVISION_DOMAIN, learner_id, sheet, row, SOURCE_COLUMN, raw_cell_sha256)}"


def _trimmed_range(value: str, start: int, end: int) -> tuple[int, int]:
    while start < end and value[start].isspace():
        start += 1
    while end > start and value[end - 1].isspace():
        end -= 1
    return start, end


def extract_source_units(raw_cell: str) -> tuple[str, str | None, list[tuple[int, int]]]:
    if "\r" in raw_cell:
        start, end = _trimmed_range(raw_cell, 0, len(raw_cell))
        return "review-required", "unsupported-carriage-return", [(start, end)]
    if "\n" not in raw_cell:
        start, end = _trimmed_range(raw_cell, 0, len(raw_cell))
        return "review-required", "no-raw-lf", [(start, end)]

    ranges: list[tuple[int, int]] = []
    start = 0
    for segment in raw_cell.split("\n"):
        end = start + len(segment)
        ranges.append(_trimmed_range(raw_cell, start, end))
        start = end + 1
    if any(start == end for start, end in ranges):
        start, end = _trimmed_range(raw_cell, 0, len(raw_cell))
        return "review-required", "empty-raw-lf-unit", [(start, end)]
    return "raw-lf", None, ranges


def _header_columns(worksheet: Any) -> dict[str, int]:
    columns: dict[str, int] = {}
    for cell in worksheet[1]:
        if isinstance(cell.value, str):
            header = cell.value.strip()
            if header == SOURCE_COLUMN and header in columns:
                raise ContractError(
                    f"sheet '{worksheet.title}' has duplicate source header '{SOURCE_COLUMN}'"
                )
            columns[header] = cell.column
    return columns


def _is_plain_text_cell(cell: Any) -> bool:
    return isinstance(cell.value, str) and cell.data_type in {"s", "inlineStr"}


def build_sidecar(
    manifest: dict[str, Any],
    workbook_path: Path,
    *,
    duplicate_discriminators: dict[str, dict[int, str]] | None = None,
) -> dict[str, Any]:
    manifest_rows = manifest.get("rows")
    if not isinstance(manifest_rows, list):
        raise ContractError("learner manifest rows must be an array")
    learner_ids: set[str] = set()
    source_coordinates: set[tuple[object, object]] = set()
    for row in manifest_rows:
        learner_id = row.get("learnerId") if isinstance(row, dict) else None
        if not isinstance(learner_id, str) or not learner_id:
            raise ContractError("learner manifest contains an invalid learner ID")
        if learner_id in learner_ids:
            raise ContractError(f"learner manifest has duplicate learner ID '{learner_id}'")
        learner_ids.add(learner_id)
        coordinate = (row.get("sourceSheet"), row.get("sourceRow"))
        if coordinate in source_coordinates:
            raise ContractError(
                f"learner manifest has duplicate source coordinate '{coordinate[0]}:{coordinate[1]}'"
            )
        source_coordinates.add(coordinate)

    workbook_sha256 = sha256_file(workbook_path)
    validate_manifest_workbook_digest(manifest, workbook_sha256)
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        records: list[dict[str, Any]] = []
        seen_phrase_ids: set[str] = set()
        duplicate_discriminators = duplicate_discriminators or {}

        for row in manifest_rows:
            sheet_name = row["sourceSheet"]
            source_row = row["sourceRow"]
            if sheet_name not in workbook.sheetnames:
                raise ContractError(f"unknown source sheet '{sheet_name}'")
            worksheet = workbook[sheet_name]
            headers = _header_columns(worksheet)
            if SOURCE_COLUMN not in headers:
                raise ContractError(f"sheet '{sheet_name}' is missing '{SOURCE_COLUMN}'")
            cell = worksheet.cell(source_row, headers[SOURCE_COLUMN])
            raw_cell = cell.value
            if raw_cell is None:
                if "example" in row:
                    raise ContractError(
                        f"source cell {sheet_name}:{source_row} is empty but manifest has example"
                    )
                continue
            if not _is_plain_text_cell(cell):
                raise ContractError(f"source cell {sheet_name}:{source_row} must be plain text")
            if not raw_cell.strip():
                raise ContractError(f"source cell {sheet_name}:{source_row} is whitespace-only")
            if normalize_example(raw_cell) != row.get("example"):
                raise ContractError(
                    f"source cell {sheet_name}:{source_row} does not match manifest example"
                )

            raw_digest = sha256_bytes(raw_cell.encode("utf-8"))
            segmentation, segmentation_reason, ranges = extract_source_units(raw_cell)
            learner_id = row["learnerId"]
            source_ref = (
                f"teacher-workbook:sha256:{workbook_sha256}"
                f"#{sheet_name}:{source_row}:{SOURCE_COLUMN}"
            )
            simplified_units = [
                unicodedata.normalize("NFC", raw_cell[start:end]) for start, end in ranges
            ]
            semantic_counts = Counter(unit.strip() for unit in simplified_units)
            record_discriminators = duplicate_discriminators.get(learner_id, {})
            used_discriminator_indices: set[int] = set()
            phrases = []
            for index, ((start, end), simplified) in enumerate(
                zip(ranges, simplified_units, strict=True)
            ):
                semantic_unit = simplified.strip()
                discriminator = ""
                if semantic_counts[semantic_unit] > 1:
                    discriminator = record_discriminators.get(index, "").strip()
                    if not discriminator:
                        raise ContractError(
                            f"learner '{learner_id}' has duplicate source unit {semantic_unit!r}; "
                            "every occurrence requires an explicit duplicate discriminator"
                        )
                    used_discriminator_indices.add(index)
                current_phrase_id = phrase_id(learner_id, simplified, discriminator)
                if current_phrase_id in seen_phrase_ids:
                    raise ContractError(f"phrase ID collision '{current_phrase_id}'")
                seen_phrase_ids.add(current_phrase_id)
                phrase = {
                    "phraseId": current_phrase_id,
                    "sourceRange": {"start": start, "end": end},
                    "simplified": simplified,
                    "fieldProvenance": {
                        "simplified": {
                            "provenance": "authored",
                            "sourceRef": source_ref,
                            "rightsRef": RIGHTS_REF,
                        }
                    },
                }
                if discriminator:
                    phrase["duplicateDiscriminator"] = discriminator
                phrases.append(phrase)
            unused_discriminator_indices = sorted(
                set(record_discriminators) - used_discriminator_indices
            )
            if unused_discriminator_indices:
                raise ContractError(
                    f"learner '{learner_id}' has unused duplicate discriminator indices "
                    f"{unused_discriminator_indices}"
                )
            source = {
                "sheet": sheet_name,
                "row": source_row,
                "column": SOURCE_COLUMN,
                "rawCell": raw_cell,
                "rawCellSha256": raw_digest,
                "sourceRevision": source_revision(
                    learner_id,
                    sheet_name,
                    source_row,
                    raw_digest,
                ),
                "segmentation": segmentation,
            }
            if segmentation_reason is not None:
                source["segmentationReason"] = segmentation_reason
            records.append(
                {
                    "learnerId": learner_id,
                    "source": source,
                    "teacherPhrases": phrases,
                }
            )

        unused_discriminator_learners = sorted(
            set(duplicate_discriminators) - {record["learnerId"] for record in records}
        )
        if unused_discriminator_learners:
            raise ContractError(
                "unused duplicate discriminator learner IDs: "
                + ", ".join(unused_discriminator_learners)
            )
    finally:
        workbook.close()

    return {
        "schemaVersion": 1,
        "contractId": CONTRACT_ID,
        "base": {
            "learnerManifestSemanticSha256": manifest_semantic_sha256(manifest),
            "workbookSha256": workbook_sha256,
        },
        "records": records,
    }


def validate_sidecar(
    sidecar: dict[str, Any],
    manifest: dict[str, Any],
    workbook_path: Path,
) -> None:
    if set(sidecar) != {"schemaVersion", "contractId", "base", "records"}:
        raise ContractError("sidecar root keys do not match the frozen contract")
    schema_version = sidecar.get("schemaVersion")
    if (
        type(schema_version) is not int
        or schema_version != 1
        or sidecar.get("contractId") != CONTRACT_ID
    ):
        raise ContractError("unsupported teacher phrase sidecar contract")
    base = sidecar.get("base")
    if not isinstance(base, dict):
        raise ContractError("sidecar base must be an object")
    if set(base) != {"learnerManifestSemanticSha256", "workbookSha256"}:
        raise ContractError("sidecar base keys do not match the frozen contract")
    expected_manifest_digest = manifest_semantic_sha256(manifest)
    if base.get("learnerManifestSemanticSha256") != expected_manifest_digest:
        raise ContractError("learner manifest semantic digest mismatch")
    expected_workbook_digest = sha256_file(workbook_path)
    validate_manifest_workbook_digest(manifest, expected_workbook_digest)
    if base.get("workbookSha256") != expected_workbook_digest:
        raise ContractError("workbook digest mismatch")

    manifest_rows = manifest.get("rows")
    if not isinstance(manifest_rows, list):
        raise ContractError("learner manifest rows must be an array")
    manifest_by_id: dict[str, dict[str, Any]] = {}
    manifest_source_keys: set[tuple[object, object]] = set()
    for row in manifest_rows:
        if not isinstance(row, dict) or not isinstance(row.get("learnerId"), str):
            raise ContractError("learner manifest contains an invalid learner ID")
        learner_id = row["learnerId"]
        if learner_id in manifest_by_id:
            raise ContractError(f"learner manifest has duplicate learner ID '{learner_id}'")
        manifest_by_id[learner_id] = row
        source_key = (row.get("sourceSheet"), row.get("sourceRow"))
        if source_key in manifest_source_keys:
            raise ContractError(
                f"learner manifest has duplicate source coordinate '{source_key[0]}:{source_key[1]}'"
            )
        manifest_source_keys.add(source_key)

    records = sidecar.get("records")
    if not isinstance(records, list):
        raise ContractError("sidecar records must be an array")
    seen_learner_ids: set[str] = set()
    seen_phrase_ids: set[str] = set()
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        for record in records:
            if not isinstance(record, dict) or not isinstance(record.get("learnerId"), str):
                raise ContractError("sidecar record has an invalid learner ID")
            if set(record) != {"learnerId", "source", "teacherPhrases"}:
                raise ContractError("sidecar record keys do not match the frozen contract")
            learner_id = record["learnerId"]
            if learner_id in seen_learner_ids:
                raise ContractError(f"sidecar has duplicate learner ID '{learner_id}'")
            seen_learner_ids.add(learner_id)
            manifest_row = manifest_by_id.get(learner_id)
            if manifest_row is None:
                raise ContractError(f"sidecar has unknown learner ID '{learner_id}'")

            source = record.get("source")
            if not isinstance(source, dict):
                raise ContractError(f"learner '{learner_id}' source must be an object")
            expected_source_keys = {
                "sheet",
                "row",
                "column",
                "rawCell",
                "rawCellSha256",
                "sourceRevision",
                "segmentation",
            }
            if source.get("segmentation") == "review-required":
                expected_source_keys.add("segmentationReason")
            if set(source) != expected_source_keys:
                raise ContractError(f"learner '{learner_id}' source keys do not match the frozen contract")
            expected_coordinate = (
                manifest_row.get("sourceSheet"),
                manifest_row.get("sourceRow"),
                SOURCE_COLUMN,
            )
            actual_coordinate = (
                source.get("sheet"),
                source.get("row"),
                source.get("column"),
            )
            if type(source.get("row")) is not int or actual_coordinate != expected_coordinate:
                raise ContractError(f"learner '{learner_id}' source coordinate mismatch")

            sheet_name, source_row, _ = expected_coordinate
            if sheet_name not in workbook.sheetnames:
                raise ContractError(f"unknown source sheet '{sheet_name}'")
            worksheet = workbook[sheet_name]
            headers = _header_columns(worksheet)
            if SOURCE_COLUMN not in headers:
                raise ContractError(f"sheet '{sheet_name}' is missing '{SOURCE_COLUMN}'")
            cell = worksheet.cell(source_row, headers[SOURCE_COLUMN])
            raw_cell = cell.value
            if not _is_plain_text_cell(cell):
                raise ContractError(f"source cell {sheet_name}:{source_row} must be plain text")
            if source.get("rawCell") != raw_cell:
                raise ContractError(f"learner '{learner_id}' rawCell does not match workbook")
            raw_digest = sha256_bytes(raw_cell.encode("utf-8"))
            if source.get("rawCellSha256") != raw_digest:
                raise ContractError(f"learner '{learner_id}' rawCellSha256 mismatch")
            expected_revision = source_revision(
                learner_id,
                sheet_name,
                source_row,
                raw_digest,
            )
            if source.get("sourceRevision") != expected_revision:
                raise ContractError(f"learner '{learner_id}' sourceRevision mismatch")
            if normalize_example(raw_cell) != manifest_row.get("example"):
                raise ContractError(f"learner '{learner_id}' rawCell does not match manifest example")

            expected_segmentation, expected_reason, expected_ranges = extract_source_units(raw_cell)
            if source.get("segmentation") != expected_segmentation:
                raise ContractError(f"learner '{learner_id}' segmentation mismatch")
            if source.get("segmentationReason") != expected_reason:
                raise ContractError(f"learner '{learner_id}' segmentationReason mismatch")

            phrases = record.get("teacherPhrases")
            if not isinstance(phrases, list) or len(phrases) != len(expected_ranges):
                raise ContractError(f"learner '{learner_id}' teacherPhrases do not cover every source unit")
            expected_units = [
                unicodedata.normalize("NFC", raw_cell[start:end])
                for start, end in expected_ranges
            ]
            semantic_counts = Counter(unit.strip() for unit in expected_units)
            for phrase_index, (phrase, expected_range, expected_unit) in enumerate(
                zip(phrases, expected_ranges, expected_units, strict=True)
            ):
                if not isinstance(phrase, dict):
                    raise ContractError(f"learner '{learner_id}' phrase {phrase_index} must be an object")
                allowed_phrase_keys = {
                    "phraseId",
                    "sourceRange",
                    "simplified",
                    "traditional",
                    "pinyin",
                    "japanese",
                    "fieldProvenance",
                    "duplicateDiscriminator",
                }
                required_phrase_keys = {
                    "phraseId",
                    "sourceRange",
                    "simplified",
                    "fieldProvenance",
                }
                if not required_phrase_keys.issubset(phrase) or not set(phrase).issubset(allowed_phrase_keys):
                    raise ContractError(
                        f"learner '{learner_id}' phrase {phrase_index} keys do not match the frozen contract"
                    )
                source_range = phrase.get("sourceRange")
                if (
                    not isinstance(source_range, dict)
                    or set(source_range) != {"start", "end"}
                    or type(source_range.get("start")) is not int
                    or type(source_range.get("end")) is not int
                ):
                    raise ContractError(f"learner '{learner_id}' phrase {phrase_index} sourceRange is malformed")
                expected_source_range = {
                    "start": expected_range[0],
                    "end": expected_range[1],
                }
                if source_range != expected_source_range:
                    raise ContractError(f"learner '{learner_id}' phrase {phrase_index} sourceRange mismatch")
                if phrase.get("simplified") != expected_unit:
                    raise ContractError(f"learner '{learner_id}' phrase {phrase_index} source unit mismatch")
                duplicate = semantic_counts[expected_unit.strip()] > 1
                discriminator = phrase.get("duplicateDiscriminator", "")
                if duplicate and (not isinstance(discriminator, str) or not discriminator.strip()):
                    raise ContractError(
                        f"learner '{learner_id}' duplicate source unit requires an explicit duplicate discriminator"
                    )
                if not duplicate and "duplicateDiscriminator" in phrase:
                    raise ContractError(
                        f"learner '{learner_id}' phrase {phrase_index} has an unexpected duplicate discriminator"
                    )
                expected_phrase_id = phrase_id(learner_id, expected_unit, discriminator)
                if phrase.get("phraseId") != expected_phrase_id:
                    raise ContractError(f"learner '{learner_id}' phrase {phrase_index} phraseId mismatch")
                if expected_phrase_id in seen_phrase_ids:
                    raise ContractError(f"phrase ID collision '{expected_phrase_id}'")
                seen_phrase_ids.add(expected_phrase_id)

                learner_fields = {
                    field
                    for field in ("simplified", "traditional", "pinyin", "japanese")
                    if field in phrase
                }
                for field in learner_fields:
                    if not isinstance(phrase[field], str) or not phrase[field].strip():
                        raise ContractError(
                            f"learner '{learner_id}' phrase {phrase_index} field '{field}' must be non-empty text"
                        )
                field_provenance = phrase.get("fieldProvenance")
                if not isinstance(field_provenance, dict) or set(field_provenance) != learner_fields:
                    raise ContractError(
                        f"learner '{learner_id}' phrase {phrase_index} fieldProvenance keys "
                        "must exactly match learner-visible fields"
                    )
                for field, evidence in field_provenance.items():
                    if not isinstance(evidence, dict):
                        raise ContractError(
                            f"learner '{learner_id}' phrase {phrase_index} '{field}' provenance must be an object"
                        )
                    if set(evidence) - {"provenance", "sourceRef", "rightsRef"}:
                        raise ContractError(
                            f"learner '{learner_id}' phrase {phrase_index} '{field}' provenance keys are invalid"
                        )
                    provenance = evidence.get("provenance")
                    if provenance not in {"authored", "generated", "verified"}:
                        raise ContractError(
                            f"learner '{learner_id}' phrase {phrase_index} '{field}' has invalid provenance"
                        )
                    for evidence_field in ("sourceRef", "rightsRef"):
                        value = evidence.get(evidence_field)
                        if not isinstance(value, str) or not value.strip():
                            raise ContractError(
                                f"learner '{learner_id}' phrase {phrase_index} '{field}' "
                                f"requires non-empty {evidence_field}"
                            )
                expected_simplified_provenance = {
                    "provenance": "authored",
                    "sourceRef": (
                        f"teacher-workbook:sha256:{expected_workbook_digest}"
                        f"#{sheet_name}:{source_row}:{SOURCE_COLUMN}"
                    ),
                    "rightsRef": RIGHTS_REF,
                }
                if field_provenance.get("simplified") != expected_simplified_provenance:
                    raise ContractError(
                        f"learner '{learner_id}' phrase {phrase_index} simplified provenance "
                        "must remain bound to the workbook source and approved rights evidence"
                    )
    finally:
        workbook.close()

    expected_record_ids = {
        row["learnerId"]
        for row in manifest_rows
        if isinstance(row.get("example"), str) and row["example"].strip()
    }
    if seen_learner_ids != expected_record_ids:
        missing = sorted(expected_record_ids - seen_learner_ids)
        extra = sorted(seen_learner_ids - expected_record_ids)
        raise ContractError(f"sidecar record coverage mismatch: missing={missing}, extra={extra}")


def refresh_sidecar(
    manifest: dict[str, Any],
    workbook_path: Path,
    *,
    existing: dict[str, Any] | None = None,
    duplicate_discriminators: dict[str, dict[int, str]] | None = None,
) -> dict[str, Any]:
    """Build a fresh sidecar or preserve an exact validated candidate.

    Existing candidate data is never merged across source/base drift.  Such a
    refresh requires explicit reconciliation by a later authoring workflow.
    """
    if existing is None:
        return build_sidecar(
            manifest,
            workbook_path,
            duplicate_discriminators=duplicate_discriminators,
        )
    validate_sidecar(existing, manifest, workbook_path)
    return copy.deepcopy(existing)
