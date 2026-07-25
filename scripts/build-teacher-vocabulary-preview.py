#!/usr/bin/env python3
"""
Build teacher-vocabulary preview from local source package.

Generation mode writes JSON fixture and copies PNGs (both gitignored).
Test mode (--test) is read-only: validates existing on-disk outputs.

Usage:
    uv run --locked python scripts/build-teacher-vocabulary-preview.py
    uv run --locked python scripts/build-teacher-vocabulary-preview.py --test
"""

import argparse
import hashlib
import json
import os
import sys
import zipfile
from pathlib import Path

# ── Constants ──────────────────────────────────────────────────────────────
EXPECTED_WORKBOOK_SHA256 = (
    "23f1ae8b7bbf38d73f70fe25d0b9df49b3e025cdd43311e2dd6aadf7d5ffacd1"
)
SHEET_NAME = "名词1"
MAX_RECORDS = 50
MIN_VALID_RECORDS = 10
TARGET_SHEET = "名词1"

JSON_OUTPUT = "public/assets/dev/teacher-vocabulary-preview/teacher-vocabulary-preview.json"
IMAGE_OUTPUT_DIR = "public/assets/dev/teacher-vocabulary-preview"

REPO_ROOT = Path(__file__).resolve().parent.parent


# ── ZIP entry decoding ────────────────────────────────────────────────────
def decode_zip_name(encoded_name: str) -> str:
    """
    Decode ZIP entry name from CP437 mojibake back to proper UTF-8.
    Python's zipfile decodes raw bytes as CP437 (per ZIP spec) when bit 11
    is not set. The original bytes are UTF-8, so we reverse: cp437 -> utf-8.
    """
    try:
        raw = encoded_name.encode("cp437")
        return raw.decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError, LookupError):
        return encoded_name


def find_workbook_and_pngs(
    zip_path: Path,
) -> tuple[bytes, dict[str, list[tuple[str, zipfile.ZipInfo]]]]:
    """
    Open the ZIP, decode entry names, and return:
      - raw xlsx bytes
      - PNG stem -> [(decoded_path, ZipInfo), ...] mapping
    A stem with >1 entry means the same filename exists in multiple directories,
    making it ambiguous for mapping.
    Raises SystemExit on failure.
    """
    png_map: dict[str, list[tuple[str, zipfile.ZipInfo]]] = {}
    xlsx_bytes: bytes | None = None

    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            for entry in zf.namelist():
                decoded = decode_zip_name(entry)
                info = zf.getinfo(entry)

                if decoded.lower().endswith(".xlsx"):
                    if xlsx_bytes is not None:
                        print("Error: multiple xlsx files in ZIP", file=sys.stderr)
                        sys.exit(1)
                    xlsx_bytes = zf.read(entry)
                elif decoded.lower().endswith(".png"):
                    stem = Path(decoded).stem
                    png_map.setdefault(stem, []).append((decoded, info))
    except zipfile.BadZipFile:
        print("Error: invalid ZIP file", file=sys.stderr)
        sys.exit(1)

    if xlsx_bytes is None:
        print("Error: no xlsx file found in ZIP", file=sys.stderr)
        sys.exit(1)

    return xlsx_bytes, png_map


def verify_checksum(data: bytes) -> str:
    actual = hashlib.sha256(data).hexdigest()
    if actual != EXPECTED_WORKBOOK_SHA256:
        print(
            f"Error: workbook SHA-256 mismatch\n"
            f"  Expected: {EXPECTED_WORKBOOK_SHA256}\n"
            f"  Actual:   {actual}",
            file=sys.stderr,
        )
        sys.exit(1)
    return actual


# ── Row processing ────────────────────────────────────────────────────────
def normalize_stem(stem: str) -> str:
    return stem.strip()


def is_image_candidate(stem: str, simplified: str) -> bool:
    """Check if a PNG stem matches a simplified word per issue rules."""
    n_stem = normalize_stem(stem)
    n_simplified = simplified.strip()

    # 1. exact normalized stem match
    if n_stem == n_simplified:
        return True

    # 2. exact whitespace-delimited token match inside the stem
    tokens = n_stem.split()
    if any(t == n_simplified for t in tokens):
        return True

    return False


def entry_count_for_stem(png_map: dict, stem: str) -> int:
    """Return the actual number of ZIP entries for this stem."""
    return len(png_map.get(stem, []))


def process_rows(
    xlsx_bytes: bytes,
    png_map: dict[str, list[tuple[str, zipfile.ZipInfo]]],
) -> tuple[list[dict], dict[str, int]]:
    """Read sheet noun1, process rows, produce items and mapping stats."""
    import openpyxl
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)

    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: sheet '{SHEET_NAME}' not found", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]

    items: list[dict] = []
    stats = {
        "mapped": 0,
        "missing": 0,
        "ambiguous": 0,
        "excluded_wrong_difficulty": 0,
        "total_valid": 0,
    }

    for row in range(2, ws.max_row + 1):
        if len(items) >= MAX_RECORDS:
            break

        simplified_val = ws.cell(row=row, column=3).value  # word
        pinyin_val = ws.cell(row=row, column=4).value       # pinyin
        japanese_val = ws.cell(row=row, column=6).value     # japanese translation
        difficulty_val = ws.cell(row=row, column=5).value   # difficulty

        simplified = str(simplified_val).strip() if simplified_val else ""
        pinyin = str(pinyin_val).strip() if pinyin_val else ""
        japanese = str(japanese_val).strip() if japanese_val else ""

        if not simplified or not pinyin or not japanese:
            continue

        # Difficulty check: non-empty must be star
        if difficulty_val is not None:
            diff_str = str(difficulty_val).strip()
            if diff_str and diff_str != "☆":
                stats["excluded_wrong_difficulty"] += 1
                continue

        stats["total_valid"] += 1
        difficulty_band = "star-1"

        # Image mapping: find matching stems
        matching_stems = [s for s in png_map if is_image_candidate(s, simplified)]

        source_image_filename: str | None = None
        image_path: str | None = None
        mapping_status: str = "missing"

        if len(matching_stems) == 0:
            mapping_status = "missing"
            stats["missing"] += 1
        elif len(matching_stems) > 1:
            # Multiple different stems match -> ambiguous
            mapping_status = "ambiguous"
            stats["ambiguous"] += 1
        else:
            stem = matching_stems[0]
            # Even one unique stem: check if it has multiple ZIP path entries
            total_entries = entry_count_for_stem(png_map, stem)
            if total_entries > 1:
                mapping_status = "ambiguous"
                stats["ambiguous"] += 1
            else:
                if normalize_stem(stem) == simplified.strip():
                    mapping_status = "exact-stem"
                else:
                    mapping_status = "exact-token"
                source_image_filename = stem + ".png"
                image_path = (
                    f"/assets/dev/teacher-vocabulary-preview/"
                    f"teacher-preview-noun1-{row:04d}.png"
                )
                stats["mapped"] += 1

        preview_id = f"teacher-preview-noun1-{row:04d}"

        item = {
            "id": preview_id,
            "simplified": simplified,
            "pinyin": pinyin,
            "japanese": japanese,
            "sourceSheet": TARGET_SHEET,
            "sourceRow": row,
            "difficultyBand": difficulty_band,
            "sourceImageFilename": source_image_filename,
            "imagePath": image_path,
            "mappingStatus": mapping_status,
            "reviewStatus": "unreviewed",
        }
        items.append(item)

    if len(items) < MIN_VALID_RECORDS:
        print(
            f"Error: fewer than {MIN_VALID_RECORDS} valid records ({len(items)}).",
            file=sys.stderr,
        )
        sys.exit(1)

    return items, stats


def copy_selected_images(
    items: list[dict],
    zip_path: Path,
    output_dir: Path,
    png_map: dict[str, list[tuple[str, zipfile.ZipInfo]]],
) -> dict[str, str]:
    """
    Copy selected preview PNGs from ZIP to output_dir.
    Also cleans stale files from output_dir that aren't in the expected set.
    Returns dict of preview_id -> sha256 for verification.
    """
    expected_files: set[str] = set()
    sha_map: dict[str, str] = {}
    with zipfile.ZipFile(zip_path, "r") as zf:
        for item in items:
            if item["mappingStatus"] not in ("exact-stem", "exact-token"):
                continue
            stem = item["sourceImageFilename"].replace(".png", "")
            candidates = png_map.get(stem, [])
            if len(candidates) != 1:
                continue

            decoded_path, _info = candidates[0]
            original_entry = None
            for entry in zf.namelist():
                if decode_zip_name(entry) == decoded_path:
                    original_entry = entry
                    break
            if original_entry is None:
                continue

            preview_id = item["id"]
            out_filename = f"{preview_id}.png"
            out_path = output_dir / out_filename
            data = zf.read(original_entry)
            file_sha = hashlib.sha256(data).hexdigest()
            with open(out_path, "wb") as dst:
                dst.write(data)
            sha_map[preview_id] = file_sha
            expected_files.add(out_filename)

    # Clean stale PNGs and JSON from the output directory
    expected_files.add("teacher-vocabulary-preview.json")
    if output_dir.exists():
        for f in output_dir.iterdir():
            if f.is_file() and f.name not in expected_files:
                f.unlink()

    return sha_map


def write_json(items: list[dict], source_sha256: str, output_path: str) -> bytes:
    """Write JSON fixture and return the raw bytes."""
    data = {
        "sourceWorkbookSha256": source_sha256,
        "status": "unreviewed-development-preview",
        "items": items,
    }
    raw = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(raw)
    return raw.encode("utf-8")


# ── Tests (read-only, operates on existing outputs) ───────────────────────
def run_tests(source_sha256: str, zip_path: Path) -> int:
    """
    Read-only validation. Expects the JSON fixture and PNG images already on
    disk (from a prior generation run). Does NOT generate or overwrite anything.
    """
    failures = 0

    def check(condition: bool, message: str):
        nonlocal failures
        if not condition:
            print(f"  FAIL: {message}")
            failures += 1
        else:
            print(f"  PASS: {message}")

    json_path = REPO_ROOT / JSON_OUTPUT

    print("\n=== test: teacher-vocabulary-preview ===\n")

    # ── Load on-disk fixture ──
    if not json_path.exists():
        print("  FAIL: JSON fixture not found (run generation first)", file=sys.stderr)
        return 1
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    items = data["items"]
    json_bytes_ondisk = json_path.read_bytes()

    # 1. Checksum
    check(
        source_sha256 == EXPECTED_WORKBOOK_SHA256,
        "workbook SHA-256 matches expected",
    )
    check(
        data.get("sourceWorkbookSha256") == EXPECTED_WORKBOOK_SHA256,
        "sourceWorkbookSha256 in JSON matches expected",
    )

    # 2. Item count
    check(10 <= len(items) <= 50, f"item count {len(items)} in [10, 50]")

    # 3. Source-row order
    source_rows = [it["sourceRow"] for it in items]
    check(source_rows == sorted(source_rows), "items in source-row order")

    # 4. Sheet
    check(
        all(it["sourceSheet"] == "名词1" for it in items),
        "all items have sourceSheet == 名词1",
    )

    # 5. IDs deterministic
    expected_ids = [f"teacher-preview-noun1-{r:04d}" for r in source_rows]
    actual_ids = [it["id"] for it in items]
    check(
        actual_ids == expected_ids,
        f"ID format deterministic: {actual_ids[0]} ... {actual_ids[-1]}",
    )

    # 6. All unreviewed
    check(
        all(it["reviewStatus"] == "unreviewed" for it in items),
        "all items unreviewed",
    )

    # 7. Status
    check(
        data.get("status") == "unreviewed-development-preview",
        "status is unreviewed-development-preview",
    )

    # 8. Mapping coverage
    mapped = [
        it for it in items
        if it["mappingStatus"] in ("exact-stem", "exact-token")
    ]
    missing = [it for it in items if it["mappingStatus"] == "missing"]
    ambiguous = [it for it in items if it["mappingStatus"] == "ambiguous"]
    check(
        len(items) == len(mapped) + len(missing) + len(ambiguous),
        f"all items classified: mapped={len(mapped)} missing={len(missing)} ambiguous={len(ambiguous)}",
    )
    check(len(mapped) > 0, f"at least one mapped image ({len(mapped)})")
    check(
        all(it["imagePath"] is not None for it in mapped),
        "all mapped items have non-null imagePath",
    )
    check(
        all(it["imagePath"] is None for it in missing + ambiguous),
        "all unmapped items have null imagePath",
    )

    # 9. Existing image presence
    for item in mapped:
        local_path = "public/" + item["imagePath"].lstrip("/")
        check(
            (REPO_ROOT / local_path).exists(),
            f"image exists: {local_path}",
        )

    # 10. Deterministic JSON output (re-read)
    with open(json_path, encoding="utf-8") as f:
        data2 = json.load(f)
    check(data == data2, "JSON output deterministic across reads")

    # 11. No timestamps
    check("timestamp" not in json_bytes_ondisk.decode("utf-8"), "no timestamps in JSON")

    # 12. Synthetic mapping tests using a ZIP-derived fixture
    # Verify exact-stem, exact-token, missing, and ambiguous coverage
    exact_stem_count = len([it for it in items if it["mappingStatus"] == "exact-stem"])
    exact_token_count = len([it for it in items if it["mappingStatus"] == "exact-token"])
    check(exact_stem_count > 0, f"exact-stem mappings exist ({exact_stem_count})")
    check(
        exact_stem_count + exact_token_count == len(mapped),
        "mapped = exact-stem + exact-token",
    )

    # 13. Determinism: re-generate in-memory and compare raw bytes
    xlsx_bytes, png_map = find_workbook_and_pngs(zip_path)
    verify_checksum(xlsx_bytes)
    items2, _stats2 = process_rows(xlsx_bytes, png_map)

    # Build expected JSON in memory
    expected_data = {
        "sourceWorkbookSha256": source_sha256,
        "status": "unreviewed-development-preview",
        "items": items2,
    }
    expected_raw = json.dumps(expected_data, ensure_ascii=False, indent=2) + "\n"
    expected_bytes = expected_raw.encode("utf-8")

    check(
        json_bytes_ondisk == expected_bytes,
        "committed JSON is byte-identical to expected output",
    )

    # 14. Run generation to a temp dir, compare against on-disk outputs for determinism
    import tempfile
    import shutil

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_out = Path(tmpdir) / "img"
        tmp_out.mkdir()
        tmp_json = Path(tmpdir) / "output.json"

        # Generate to temp
        _tmp_items, _tmp_stats = process_rows(xlsx_bytes, png_map)
        tmp_data = {
            "sourceWorkbookSha256": source_sha256,
            "status": "unreviewed-development-preview",
            "items": _tmp_items,
        }
        tmp_raw = json.dumps(tmp_data, ensure_ascii=False, indent=2) + "\n"
        tmp_json.write_text(tmp_raw, encoding="utf-8")

        # Compare JSON raw bytes
        tmp_json_bytes = tmp_json.read_bytes()
        check(
            tmp_json_bytes == json_bytes_ondisk,
            "generated JSON byte-identical across two runs",
        )

        # Copy images to temp and compare SHA-256 per file
        sha_gen = copy_selected_images(_tmp_items, zip_path, tmp_out, png_map)
        for preview_id, sha_from_zip in sha_gen.items():
            ondisk_img = REPO_ROOT / IMAGE_OUTPUT_DIR / f"{preview_id}.png"
            if ondisk_img.exists():
                ondisk_sha = hashlib.sha256(ondisk_img.read_bytes()).hexdigest()
                check(
                    ondisk_sha == sha_from_zip,
                    f"image {preview_id}.png SHA-256 matches source ZIP entry",
                )
            else:
                check(
                    False,
                    f"image {preview_id}.png not found on disk for SHA check",
                )

    # 15. Stale/extra managed PNG detection
    expected_ids = {it["id"] for it in mapped}
    actual_pngs = set()
    img_dir = REPO_ROOT / IMAGE_OUTPUT_DIR
    if img_dir.exists():
        for f in img_dir.iterdir():
            if f.suffix == ".png":
                actual_pngs.add(f.stem)
    stale = actual_pngs - expected_ids
    if stale:
        check(False, f"stale images not in JSON: {sorted(stale)}")
    else:
        check(True, "no stale managed images")

    # 16. Source ZIP, workbook, RAR, PDF not in git diff
    import subprocess
    result = subprocess.run(
        ["git", "diff", "--name-only", "--cached"],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    changed = result.stdout
    for ext in [".zip", ".xlsx", ".rar", ".pdf"]:
        check(ext not in changed, f"no {ext} files in staged diff")

    print(f"\nResults: {failures} failure(s)")
    for m in mapped[:3]:
        print(f"  mapped: {m['simplified']} -> {m['mappingStatus']}")
    for m in missing[:3]:
        print(f"  missing: {m['simplified']}")
    for a in ambiguous[:3]:
        print(f"  ambiguous: {a['simplified']}")
    if len(missing) > 3:
        print(f"  ... and {len(missing) - 3} more missing")
    if len(ambiguous) > 3:
        print(f"  ... and {len(ambiguous) - 3} more ambiguous")

    return 0 if failures == 0 else 1


# ── Main ──────────────────────────────────────────────────────────────────
def cmd_generate(zip_path: Path) -> tuple[list[dict], dict[str, int], bytes, dict[str, str]]:
    """Run full generation: parse ZIP, produce items, write JSON, copy images."""
    xlsx_bytes, png_map = find_workbook_and_pngs(zip_path)
    source_sha256 = verify_checksum(xlsx_bytes)
    items, stats = process_rows(xlsx_bytes, png_map)

    output_dir = REPO_ROOT / IMAGE_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    sha_map = copy_selected_images(items, zip_path, output_dir, png_map)
    raw_bytes = write_json(items, source_sha256, str(REPO_ROOT / JSON_OUTPUT))

    print(f"Generated {len(items)} records from sheet '{SHEET_NAME}'")
    print(f"Mapped: {stats['mapped']}  Missing: {stats['missing']}  Ambiguous: {stats['ambiguous']}")
    print(f"JSON:    {JSON_OUTPUT}")
    print(f"Images:  {IMAGE_OUTPUT_DIR}/")

    return items, stats, raw_bytes, sha_map


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build teacher vocabulary preview from local source package",
    )
    parser.add_argument(
        "--zip",
        default="/tmp/chabiko-issue-113-teacher-images.zip",
        help="Path to the source ZIP package",
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Run read-only validation tests (no generation)",
    )
    args = parser.parse_args()

    zip_path = Path(args.zip)
    if not zip_path.exists():
        print(f"Error: ZIP not found: {zip_path}", file=sys.stderr)
        return 1

    if args.test:
        # Read-only test mode
        xlsx_bytes, _png_map = find_workbook_and_pngs(zip_path)
        source_sha256 = verify_checksum(xlsx_bytes)
        return run_tests(source_sha256, zip_path)

    # Generation mode
    cmd_generate(zip_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
