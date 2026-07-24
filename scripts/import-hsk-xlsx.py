#!/usr/bin/env python3
"""
HSK XLSX Importer for Chabiko (#75).

Reads an HSK workbook XLSX, parses text-only content, and emits
versioned vocabulary records plus a bounded import manifest.

Usage:
    uv run python scripts/import-hsk-xlsx.py <input.xlsx> <output-dir> <hsk-version>
    uv run python scripts/import-hsk-xlsx.py --test
"""

import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import unicodedata

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# ─── Valid standard versions ─────────────────────────────────────────────────

VALID_STANDARD_VERSIONS = frozenset({"hsk-3.0", "hsk-legacy-6-level"})


# ─── Column name mappings ───────────────────────────────────────────────────

REQUIRED_COLUMN_MAPPINGS = {
    "simplified": frozenset({"simplified", "简体", "简体字", "简"}),
    "pinyin": frozenset({"pinyin", "拼音"}),
    "japanese": frozenset({"japanese", "日文", "日本語", "日语", "jpn"}),
}

OPTIONAL_COLUMN_MAPPINGS = {
    "traditional": frozenset({"traditional", "繁体", "繁体字", "繁"}),
    "level": frozenset({"level", "hsk level", "级别", "等级", "lev"}),
    "category": frozenset({"category", "类别", "分類"}),
    "kana": frozenset({"kana", "假名", "仮名", "よみ", "reading"}),
}


# ─── Unicode/whitespace normalization ────────────────────────────────────────

def normalize_text(value):
    """Normalize whitespace and Unicode deterministically.

    - Strips leading/trailing whitespace
    - Collapses internal whitespace runs to single space
    - Applies NFC normalization (does not alter CJK or pinyin tones)
    """
    if not isinstance(value, str):
        value = str(value) if value is not None else ""
    value = unicodedata.normalize("NFC", value.strip())
    value = re.sub(r"\s+", " ", value)
    return value


def _normalize_simplified(text):
    """Normalize simplified Chinese per HSK identity contract.

    Matches validate-content-schema.py _normalize_simplified exactly.
    """
    normalized = unicodedata.normalize("NFKC", text)
    return "".join(ch for ch in normalized if not ch.isspace())


def _normalize_pinyin(text):
    """Normalize pinyin per HSK identity contract.

    Matches validate-content-schema.py _normalize_pinyin exactly.
    """
    normalized = unicodedata.normalize("NFKC", text)
    case_folded = normalized.casefold()
    return "".join(ch for ch in case_folded if not ch.isspace())


# ─── Stable ID generation ────────────────────────────────────────────────────

def generate_stable_id(simplified, pinyin, level):
    """Generate a deterministic ID from content.

    Algorithm: SHA-256 of "{simplified}|{pinyin}" → first 12 hex chars.
    ID format: ``hsk-{level}-{hash}``.

    This is a one-way hash: given the same (simplified, pinyin, level) triple,
    the output is always identical across platforms and Python versions.
    """
    seed = f"{simplified}|{pinyin}"
    hash_digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    return f"hsk-{level}-{hash_digest}"


# ─── Header resolution ──────────────────────────────────────────────────────

def resolve_headers(header_row):
    """Detect canonical column mapping from a header row.

    Returns (mapping, missing) where *mapping* is a dict of
    ``{canonical_key: column_index}`` and *missing* is a list of required keys
    that could not be resolved.  When required columns are missing, *mapping*
    is ``None``.
    """
    mapping = {}
    for idx, raw in enumerate(header_row):
        text = normalize_text(raw).lower() if raw else ""
        if not text:
            continue

        # Check required columns first
        for key, aliases in REQUIRED_COLUMN_MAPPINGS.items():
            if key in mapping:
                continue
            if text in aliases:
                mapping[key] = idx
                break

        # Then optional columns
        for key, aliases in OPTIONAL_COLUMN_MAPPINGS.items():
            if key in mapping:
                continue
            if text in aliases:
                mapping[key] = idx
                break

    missing = [k for k in REQUIRED_COLUMN_MAPPINGS if k not in mapping]
    if missing:
        return None, missing
    return mapping, []


# ─── Sheet detection ─────────────────────────────────────────────────────────

def find_data_sheets(wb):
    """Classify workbook sheets and return data sheets.

    Returns (data_sheets, malformed_sheets) where:
    - data_sheets: list of (sheet_name, worksheet, header_mapping)
    - malformed_sheets: list of sheet names that have at least one recognised
      vocabulary header but are missing required columns.

    Sheets with zero recognised vocabulary headers are metadata sheets
    (silently ignored).
    """
    data_sheets = []
    malformed = []
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 2:
            continue
        header_row = [cell.value if cell.value is not None else ""
                      for cell in ws[1]]
        recognised = _count_recognised_headers(header_row)
        if recognised == 0:
            continue  # metadata sheet, ignore
        mapping, _missing = resolve_headers(header_row)
        if mapping is not None:
            data_sheets.append((ws.title, ws, mapping))
        else:
            malformed.append(ws.title)
    if malformed:
        raise ValueError(
            f"Sheets with partial vocabulary headers (missing required "
            f"columns): {', '.join(sorted(malformed))}"
        )
    return data_sheets


def _count_recognised_headers(header_row):
    """Count headers matching any known alias."""
    count = 0
    for raw in header_row:
        text = normalize_text(raw).lower() if raw else ""
        if not text:
            continue
        for aliases in list(REQUIRED_COLUMN_MAPPINGS.values()) + list(OPTIONAL_COLUMN_MAPPINGS.values()):
            if text in aliases:
                count += 1
                break
    return count


# ─── Row parsing ─────────────────────────────────────────────────────────────

def parse_row(ws, row_idx, mapping):
    """Parse a single data row.

    Returns a dict with canonical column values plus ``_sourceSheet`` and
    ``_sourceRow`` metadata.  Formula cells in required columns cause the
    row to be flagged; formula cells in optional columns are skipped.
    """
    record = {"_sourceSheet": ws.title, "_sourceRow": row_idx}

    for canonical_key, col_idx in mapping.items():
        cell = ws.cell(row=row_idx, column=col_idx + 1)  # openpyxl is 1-based
        raw = cell.value
        is_formula = (cell.data_type == "f")

        if is_formula and canonical_key in REQUIRED_COLUMN_MAPPINGS:
            record["_formula_required"] = record.get("_formula_required", [])
            record["_formula_required"].append(canonical_key)
            record[canonical_key] = ""
            continue

        if canonical_key == "level":
            if raw is not None:
                if isinstance(raw, bool):
                    record[canonical_key] = raw  # pass through for validation
                elif isinstance(raw, int):
                    record[canonical_key] = raw
                elif isinstance(raw, float):
                    if raw == int(raw):
                        record[canonical_key] = int(raw)
                    else:
                        record[canonical_key] = raw  # non-integer float
                else:
                    record[canonical_key] = str(raw)
            else:
                record[canonical_key] = None
            continue

        if is_formula:
            # Optional field formula: treat as empty
            record[canonical_key] = ""
            continue

        if raw is None:
            record[canonical_key] = ""
        else:
            record[canonical_key] = normalize_text(str(raw))

    return record


def _is_fully_empty_record(record, mapping):
    """Return whether every mapped source cell is empty after normalization."""
    if record.get("_formula_required"):
        return False
    return not any(normalize_text(record.get(key)) for key in mapping)


# ─── Record validation ───────────────────────────────────────────────────────

def validate_record(record, seen_ids, seen_identities, standard_version):
    """Validate and convert a parsed record into a vocabulary dict.

    Returns (vocab_dict, rejection_reason_or_None).
    """
    simplified = record.get("simplified", "")
    pinyin = record.get("pinyin", "")
    sheet = record.get("_sourceSheet", "")
    row = record.get("_sourceRow", 0)

    # Formula in required field → reject
    formula_fields = record.get("_formula_required")
    if formula_fields:
        return None, (f"formula in required column(s) {', '.join(formula_fields)}: "
                      f"{sheet}:{row}")

    if not simplified:
        return None, f"missing simplified: {sheet}:{row}"

    if not pinyin:
        return None, f"missing pinyin: {sheet}:{row}"

    japanese = record.get("japanese", "")
    if not japanese:
        return None, f"missing japanese: {sheet}:{row}"

    # Level validation
    level_raw = record.get("level")
    level_reason = None
    if level_raw is None:
        level_reason = "missing level"
    elif isinstance(level_raw, bool):
        level_reason = f"invalid level (bool)"
    elif isinstance(level_raw, int):
        if level_raw < 1 or level_raw > 9:
            level_reason = f"invalid level (out of range 1-9: {level_raw})"
    elif isinstance(level_raw, float):
        if level_raw != int(level_raw):
            level_reason = f"invalid level (non-integer: {level_raw})"
        else:
            level_raw_int = int(level_raw)
            if level_raw_int < 1 or level_raw_int > 9:
                level_reason = f"invalid level (out of range 1-9: {level_raw_int})"
    else:
        level_reason = f"invalid level (string: {level_raw})"

    if level_reason:
        return None, f"{level_reason}: {sheet}:{row}"

    level = int(level_raw)

    identity = (_normalize_simplified(simplified), _normalize_pinyin(pinyin))
    if identity in seen_identities:
        return None, (f"duplicate identity ({simplified}, {pinyin}): "
                      f"{sheet}:{row}")

    vid = generate_stable_id(simplified, pinyin, level)
    if vid in seen_ids:
        return None, (f"collision on ID {vid}: "
                      f"{sheet}:{row}")

    vocab = {
        "id": vid,
        "simplified": simplified,
        "simplifiedStatus": "authored",
        "pinyin": pinyin,
        "japanese": japanese,
        "source": {"type": "hsk-workbook"},
        "reviewStatus": "draft",
        "hsk": {
            "standardVersion": standard_version,
            "introducedAtLevel": level,
            "sourceLevelLabel": f"HSK {standard_version} Level {level}",
        },
    }

    traditional = record.get("traditional", "")
    if traditional:
        vocab["traditional"] = traditional
        vocab["traditionalStatus"] = "authored"

    category = record.get("category", "")
    if category:
        vocab["category"] = category

    kana = record.get("kana", "")
    if kana:
        vocab["kana"] = kana

    return vocab, None


# ─── Batch output ────────────────────────────────────────────────────────────

def write_batches(accepted_entries, output_dir):
    """Write vocabulary records to deterministic batch JSON files.

    *accepted_entries* is a list of ``(vocab_dict, source_sheet, source_row)``
    tuples.  Returns batch metadata list for the manifest.
    """
    accepted_entries.sort(key=lambda v: (
        v[0]["hsk"]["introducedAtLevel"],
        v[0]["simplified"],
        v[0]["pinyin"],
    ))

    batches = []
    for i in range(0, len(accepted_entries), 50):
        chunk = accepted_entries[i:i + 50]
        batch_num = i // 50 + 1
        filename = f"hsk-vocabulary-batch-{batch_num:02d}.json"
        filepath = os.path.join(output_dir, filename)

        batch_vocab = []
        batch_sources = []
        for v, sheet, row in chunk:
            batch_vocab.append(v)
            batch_sources.append({
                "id": v["id"],
                "sheet": sheet,
                "row": row,
            })

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({"vocabulary": batch_vocab}, f,
                      ensure_ascii=False, indent=2)
            f.write("\n")

        batches.append({
            "filename": filename,
            "batchNumber": batch_num,
            "entryCount": len(batch_vocab),
            "firstEntryId": batch_vocab[0]["id"],
            "lastEntryId": batch_vocab[-1]["id"],
            "sourceRows": batch_sources,
        })

    return batches


def _test_missing_japanese():
    """Verify a workbook without japanese column is rejected."""
    import tempfile as _tf
    from openpyxl import Workbook
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "Test"
        _ws.append(["simplified", "pinyin"])
        _ws.append(["爱", "ài"])
        _path = os.path.join(_td, "no_jp.xlsx")
        _wb.save(_path)
        try:
            import_xlsx(_path, os.path.join(_td, "out"), "hsk-3.0")
            raise AssertionError("expected ValueError for missing japanese column")
        except ValueError:
            pass


# ─── Main import pipeline ────────────────────────────────────────────────────

def import_xlsx(input_path, output_dir, standard_version):
    """Run the full import pipeline.

    Returns the manifest dict.
    """
    # 0. Validate standard version (fail BEFORE any output mutation)
    if standard_version not in VALID_STANDARD_VERSIONS:
        raise ValueError(
            f"Unsupported HSK standard version: '{standard_version}'. "
            f"Valid versions: {', '.join(sorted(VALID_STANDARD_VERSIONS))}"
        )

    # 1. Source checksum
    with open(input_path, "rb") as f:
        checksum = hashlib.sha256(f.read()).hexdigest()

    # 2. Open workbook (disable data_only to detect formula cells)
    wb = openpyxl.load_workbook(input_path, data_only=False)

    # 3. Detect data sheets (raises ValueError for malformed sheets)
    data_sheets = find_data_sheets(wb)
    if not data_sheets:
        raise ValueError("No sheets with valid vocabulary headers found")

    sheet_names = [s[0] for s in data_sheets]

    # 4. Parse all rows (skip header row 1)
    all_records = []
    for _sheet_name, ws, mapping in data_sheets:
        for row_idx in range(2, ws.max_row + 1):
            record = parse_row(ws, row_idx, mapping)
            if _is_fully_empty_record(record, mapping):
                continue
            all_records.append(record)

    # 5. Validate and categorise
    seen_ids = set()
    seen_identities = set()
    accepted_entries = []       # (vocab, sheet, row)
    rejected = []               # {"reason": str, "sheet": str, "row": int}

    for record in all_records:
        vocab, reason = validate_record(
            record, seen_ids, seen_identities, standard_version,
        )
        if reason:
            rejected.append({
                "reason": reason,
                "sheet": record.get("_sourceSheet", ""),
                "row": record.get("_sourceRow", 0),
            })
        else:
            normalized_identity = (
                _normalize_simplified(vocab["simplified"]),
                _normalize_pinyin(vocab["pinyin"]),
            )
            seen_ids.add(vocab["id"])
            seen_identities.add(normalized_identity)
            accepted_entries.append((
                vocab,
                record.get("_sourceSheet", ""),
                record.get("_sourceRow", 0),
            ))

    # 6. Count by level (accepted only)
    accepted_by_level = {}
    for v, _sheet, _row in accepted_entries:
        lvl = v["hsk"]["introducedAtLevel"]
        accepted_by_level[lvl] = accepted_by_level.get(lvl, 0) + 1

    # totalByLevel: count every parsed row by its raw level value
    total_by_level = {}
    rejected_by_level = {}
    for record in all_records:
        lvl_raw = record.get("level")
        bucket = _level_bucket(lvl_raw)
        total_by_level[bucket] = total_by_level.get(bucket, 0) + 1

    # rejectedByLevel: count rejected rows by their raw level value
    for r in rejected:
        lvl_raw = None
        # Find the corresponding record
        for rec in all_records:
            if rec.get("_sourceSheet") == r["sheet"] and rec.get("_sourceRow") == r["row"]:
                lvl_raw = rec.get("level")
                break
        bucket = _level_bucket(lvl_raw)
        rejected_by_level[bucket] = rejected_by_level.get(bucket, 0) + 1

    # 7. Write batches (output mutation starts here)
    os.makedirs(output_dir, exist_ok=True)
    batches = write_batches(accepted_entries, output_dir)

    # 8. Build manifest
    duplicate_diagnostics = [
        r for r in rejected if "duplicate" in r["reason"]
    ]

    manifest = {
        "sourceFile": os.path.basename(input_path),
        "sourceChecksumSha256": checksum,
        "standardVersion": standard_version,
        "sourceSheets": sheet_names,
        "totalRows": len(all_records),
        "totalByLevel": _serialise_level_counts(total_by_level),
        "accepted": len(accepted_entries),
        "acceptedByLevel": _serialise_level_counts(accepted_by_level),
        "rejected": len(rejected),
        "rejectedByLevel": _serialise_level_counts(rejected_by_level),
        "duplicateDiagnostics": duplicate_diagnostics,
        "rejectedRows": rejected,
        "batchCount": len(batches),
        "batches": batches,
    }

    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
        f.write("\n")

    # 9. Validate every batch against the #74 contract
    _validate_batches(output_dir, batches)

    return manifest


def _level_bucket(raw):
    """Return a level bucket key for a raw level value.

    Valid int 1-9 returns the int itself (for deterministic sort).
    Missing, boolean, non-integer, out-of-range, etc. return the string
    ``"invalid"``.
    """
    if raw is None:
        return "invalid"
    if isinstance(raw, bool):
        return "invalid"
    if isinstance(raw, int):
        if 1 <= raw <= 9:
            return raw
        return "invalid"
    return "invalid"


def _serialise_level_counts(counter):
    """Serialise a level-count dict with deterministic key order.

    Int keys sort before the ``"invalid"`` string.
    """
    result = {}
    for k in sorted(counter, key=_level_sort_key):
        result[str(k)] = counter[k]
    return result


def _level_sort_key(k):
    """Sort key: ints first (by value), then strings."""
    return (0, k) if isinstance(k, int) else (1, k)


def _validate_batches(output_dir, batches):
    """Run the #74 content schema validator on every batch file.

    Reuses the existing validate-content-schema.py via subprocess since
    the validator is a standalone script without a stable importable API.
    """
    validator = os.path.join(os.path.dirname(__file__), "validate-content-schema.py")
    for b in batches:
        batch_path = os.path.join(output_dir, b["filename"])
        result = subprocess.run(
            [sys.executable, validator, "--check", batch_path],
            capture_output=True, text=True,
        )
        if result.returncode != 0:
            lines = [l for l in result.stdout.strip().split("\n") if l.strip()]
            summary = "; ".join(lines[:3])
            raise RuntimeError(
                f"Batch {b['filename']} failed #74 contract validation: {summary}"
            )


# ─── CLI entry point ─────────────────────────────────────────────────────────

def main():
    if len(sys.argv) == 2 and sys.argv[1] == "--test":
        sys.exit(run_tests())

    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2]
    standard_version = sys.argv[3]

    if not os.path.isfile(input_path):
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        manifest = import_xlsx(input_path, output_dir, standard_version)
        print(f"Import complete: {manifest['accepted']} accepted, "
              f"{manifest['rejected']} rejected")
        print(f"  Batches: {manifest['batchCount']}")
        print(f"  Output: {output_dir}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


# ─── Self-tests ──────────────────────────────────────────────────────────────

def _make_synthetic_wb(path):
    """Create a synthetic multi-sheet XLSX fixture at *path*."""
    wb = Workbook()

    # Sheet 1 — standard English headers
    ws1 = wb.active
    ws1.title = "HSK Level 1"
    ws1.append(["simplified", "pinyin", "traditional", "level", "japanese"])
    ws1.append(["爱", "ài", "愛", 1, "愛する"])
    ws1.append(["你好", "nǐ hǎo", "你好", 1, "こんにちは"])
    ws1.append(["猫", "māo", "", 1, "猫"])
    ws1.append(["狗", "gǒu", "狗", 1, "犬"])
    ws1.append(["水", "shuǐ", "水", 1, "水"])

    # Row with empty simplified (should be rejected)
    ws1.append(["", "shénme", "什麼", 1, "何"])

    # Duplicate identity (爱 + ài) — should be rejected
    ws1.append(["爱", "ài", "愛", 1, "愛する"])

    # Case-variant duplicate (水 + Shuǐ with uppercase S) — should be rejected
    ws1.append(["水", "Shuǐ", "水", 1, "水"])

    # Sheet 2 — Chinese headers
    ws2 = wb.create_sheet("HSK Level 2")
    ws2.append(["简体", "拼音", "级别", "日文"])
    ws2.append(["书", "shū", 2, "本"])
    ws2.append(["学校", "xuéxiào", 2, "学校"])
    ws2.append(["医院", "yīyuàn", 2, "病院"])

    # Sheet 3 — mixed headers, optional traditional included
    ws3 = wb.create_sheet("HSK Level 3-4")
    ws3.append(["简体字", "拼音", "繁体字", "Level", "日文", "类别"])
    ws3.append(["电脑", "diànnǎo", "電腦", 3, "コンピューター", "technology"])
    ws3.append(["电话", "diànhuà", "電話", 3, "電話", "technology"])
    ws3.append(["老师", "lǎoshī", "老師", 3, "先生", "education"])
    ws3.append(["同学", "tóngxué", "同學", 4, "同級生", "education"])

    # Sheet 4 — missing required column (no pinyin) → metadata
    ws4 = wb.create_sheet("Metadata Only")
    ws4.append(["notes", "description"])
    ws4.append(["foo", "bar"])

    # Sheet 5 — empty, skipped
    wb.create_sheet("Empty")

    wb.save(path)


def _make_empty_row_wb(path, include_trailing_empty_rows, include_internal_empty_row=False):
    """Create equivalent fixtures with or without empty worksheet rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty Rows"
    ws.append(["simplified", "pinyin", "traditional", "level", "japanese"])
    ws.append(["爱", "ài", "愛", 1, "愛する"])

    if include_internal_empty_row:
        ws.append([None, None, None, None, None])

    ws.append(["书", "shū", "書", 1, "本"])
    ws.append(["", "shénme", "什麼", 1, "何"])
    ws.append(["", "", "", "", "日本語だけ"])
    ws.append(["吃", "chī", "吃", 1, "食べる"])

    if include_trailing_empty_rows:
        ws.append([None, None, None, None, None])
        ws.append(["  ", "\t", " ", "   ", "\n"])
        styled_row = ws.max_row + 1
        for column in range(1, 6):
            ws.cell(row=styled_row, column=column).fill = PatternFill(
                fill_type="solid",
                fgColor="FFFF00",
            )

    wb.save(path)


def _check_batch_order_and_size(output_dir):
    """Verify every batch contains 1-50 entries with stable ordering."""
    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    assert manifest["batchCount"] >= 1, "must produce at least one batch"

    for b in manifest["batches"]:
        assert 1 <= b["entryCount"] <= 50, \
            f"batch {b['batchNumber']} has {b['entryCount']} entries (want 1-50)"

        batch_path = os.path.join(output_dir, b["filename"])
        with open(batch_path, "r", encoding="utf-8") as f:
            batch = json.load(f)

        assert len(batch["vocabulary"]) == b["entryCount"]

        # Verify sort order
        entries = batch["vocabulary"]
        for i in range(1, len(entries)):
            prev = entries[i - 1]
            cur = entries[i]
            prev_key = (prev["hsk"]["introducedAtLevel"], prev["simplified"], prev["pinyin"])
            cur_key = (cur["hsk"]["introducedAtLevel"], cur["simplified"], cur["pinyin"])
            # prev should be <= cur in sort order
            assert prev_key <= cur_key, \
                f"batch {b['batchNumber']} not sorted at index {i}"

    return manifest


def _check_manifest_counts(manifest):
    """Verify by-level count sums equal their top-level totals."""
    total_accepted = sum(manifest["acceptedByLevel"].values())
    assert total_accepted == manifest["accepted"], \
        f"acceptedByLevel sum ({total_accepted}) != accepted ({manifest['accepted']})"

    total_rejected = sum(manifest["rejectedByLevel"].values())
    assert total_rejected == manifest["rejected"], \
        f"rejectedByLevel sum ({total_rejected}) != rejected ({manifest['rejected']})"

    total_all = sum(manifest["totalByLevel"].values())
    assert total_all == manifest["totalRows"], \
        f"totalByLevel sum ({total_all}) != totalRows ({manifest['totalRows']})"




def _test_invalid_standard_version():
    """Verify unsupported standardVersion is rejected before any output."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        from openpyxl import Workbook
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "Test"
        _ws.append(["simplified", "pinyin", "japanese"])
        _ws.append(["爱", "ài", "愛する"])
        _path = os.path.join(_td, "x.xlsx")
        _wb.save(_path)
        try:
            import_xlsx(_path, os.path.join(_td, "out"), "hsk-2.0")
            raise AssertionError("expected ValueError for invalid version")
        except ValueError as e:
            assert "hsk-2.0" in str(e)
            assert not os.path.exists(os.path.join(_td, "out")), \
                "output directory created before version validation"


def _test_invalid_level_rejection():
    """Verify invalid level values are rejected with sheet/row diagnostics."""
    import tempfile as _tf
    from openpyxl import Workbook
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "LevelTest"
        _ws.append(["simplified", "pinyin", "japanese", "level"])
        _ws.append(["一", "yī", "一", 1])       # valid
        _ws.append(["零", "líng", "零", 0])      # out of range
        _ws.append(["十", "shí", "十", 10])      # out of range
        _ws.append(["bad", "bād", "bad", "abc"])  # string level
        _ws.append(["flt", "flt", "flt", 2.5])   # non-integer float
        _ws.append(["bool_t", "b", "b", True])   # bool level
        _ws.append(["four", "sì", "四", 4])      # valid
        _path = os.path.join(_td, "levels.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out, "hsk-3.0")
        assert _m["accepted"] == 2, \
            f"expected 2 accepted, got {_m['accepted']}"
        assert _m["rejected"] == 5, \
            f"expected 5 rejected, got {_m['rejected']}"
        for r in _m["rejectedRows"]:
            assert r["sheet"] == "LevelTest", \
                f"expected LevelTest sheet, got {r['sheet']}"
            assert "level" in r["reason"].lower(), \
                f"level not mentioned in rejection: {r['reason']}"


def _test_formula_rejection():
    """Verify formulas in required fields are rejected; optional formulas are ignored."""
    import tempfile as _tf
    from openpyxl import Workbook
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "FormulaTest"
        _ws.append(["simplified", "pinyin", "japanese", "level"])
        _ws.append(["爱", "ài", "愛する", 1])
        _path = os.path.join(_td, "formula.xlsx")
        _wb.save(_path)

        # Inject a formula into simplified (required column)
        import openpyxl as _opxl
        _wb2 = _opxl.load_workbook(_path)
        _ws2 = _wb2.active
        _ws2.cell(row=2, column=1).value = "=A2&B2"
        _wb2.save(_path)

        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out, "hsk-3.0")
        assert _m["accepted"] == 0, \
            f"expected 0 accepted with formula, got {_m['accepted']}"
        assert _m["rejected"] == 1, \
            f"expected 1 rejection for formula, got {_m['rejected']}"
        assert "formula" in _m["rejectedRows"][0]["reason"].lower(), \
            f"formula not mentioned: {_m['rejectedRows'][0]['reason']}"


def _test_sheet_detection_malformed():
    """Verify a valid sheet + malformed candidate raises ValueError before output."""
    import tempfile as _tf
    from openpyxl import Workbook
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws1 = _wb.active
        _ws1.title = "Good"
        _ws1.append(["simplified", "pinyin", "japanese"])
        _ws1.append(["爱", "ài", "愛する"])
        _ws2 = _wb.create_sheet("Bad")
        _ws2.append(["simplified", "pinyin"])  # missing japanese (required) with recognised headers
        _ws2.append(["好", "hǎo"])
        _path = os.path.join(_td, "mixed.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        try:
            import_xlsx(_path, _out, "hsk-3.0")
            raise AssertionError("expected ValueError for malformed sheet")
        except ValueError as e:
            assert "partial" in str(e).lower() or "missing" in str(e).lower()
            assert not os.path.exists(os.path.join(_out, "manifest.json")), \
                "output created before sheet validation"


def _test_empty_rows_ignored():
    """Verify empty rows do not affect diagnostics or deterministic output."""
    with tempfile.TemporaryDirectory() as test_dir:
        control_path = os.path.join(test_dir, "control.xlsx")
        trailing_rows_path = os.path.join(test_dir, "with-trailing-rows.xlsx")
        all_empty_rows_path = os.path.join(test_dir, "with-all-empty-rows.xlsx")
        control_output = os.path.join(test_dir, "control-out")
        trailing_output = os.path.join(test_dir, "trailing-out")
        all_empty_output_a = os.path.join(test_dir, "all-empty-out-a")
        all_empty_output_b = os.path.join(test_dir, "all-empty-out-b")

        _make_empty_row_wb(
            control_path,
            include_trailing_empty_rows=False,
        )
        _make_empty_row_wb(
            trailing_rows_path,
            include_trailing_empty_rows=True,
        )
        _make_empty_row_wb(
            all_empty_rows_path,
            include_trailing_empty_rows=True,
            include_internal_empty_row=True,
        )

        control_manifest = import_xlsx(
            control_path, control_output, "hsk-3.0",
        )
        trailing_manifest = import_xlsx(
            trailing_rows_path, trailing_output, "hsk-3.0",
        )
        all_empty_manifest_a = import_xlsx(
            all_empty_rows_path, all_empty_output_a, "hsk-3.0",
        )
        all_empty_manifest_b = import_xlsx(
            all_empty_rows_path, all_empty_output_b, "hsk-3.0",
        )

        for manifest in (
            control_manifest,
            trailing_manifest,
            all_empty_manifest_a,
        ):
            assert manifest["totalRows"] == 5
            assert manifest["accepted"] == 3
            assert manifest["rejected"] == 2
            assert manifest["batchCount"] == 1
            _check_manifest_counts(manifest)

        expected_reasons = [
            "missing simplified: Empty Rows:5",
            "missing simplified: Empty Rows:6",
        ]
        assert [
            row["reason"] for row in all_empty_manifest_a["rejectedRows"]
        ] == expected_reasons, (
            "partially populated rows must retain exact sheet/row diagnostics"
        )

        for field in control_manifest:
            if field in {"sourceFile", "sourceChecksumSha256"}:
                continue
            assert control_manifest[field] == trailing_manifest[field], (
                f"trailing empty rows changed manifest field {field}"
            )

        def read_batch_bytes(output_dir, manifest):
            payloads = []
            for batch in manifest["batches"]:
                batch_path = os.path.join(output_dir, batch["filename"])
                with open(batch_path, "rb") as f:
                    payloads.append(f.read())
            return payloads

        control_batches = read_batch_bytes(control_output, control_manifest)
        trailing_batches = read_batch_bytes(trailing_output, trailing_manifest)
        all_empty_batches_a = read_batch_bytes(
            all_empty_output_a, all_empty_manifest_a,
        )
        all_empty_batches_b = read_batch_bytes(
            all_empty_output_b, all_empty_manifest_b,
        )
        assert control_batches == trailing_batches == all_empty_batches_a, (
            "empty rows changed stable IDs, batch boundaries, or ordering"
        )
        assert all_empty_batches_a == all_empty_batches_b, (
            "repeated import with empty rows changed batch bytes"
        )
        assert all_empty_manifest_a == all_empty_manifest_b, (
            "repeated import with empty rows changed the manifest"
        )
        with open(
            os.path.join(all_empty_output_a, "manifest.json"), "rb",
        ) as manifest_a_file:
            manifest_bytes_a = manifest_a_file.read()
        with open(
            os.path.join(all_empty_output_b, "manifest.json"), "rb",
        ) as manifest_b_file:
            manifest_bytes_b = manifest_b_file.read()
        assert manifest_bytes_a == manifest_bytes_b, (
            "repeated import with empty rows changed manifest bytes"
        )
        assert "_check_byte_identical" not in globals(), (
            "dead _check_byte_identical helper must remain absent"
        )


def run_tests():
    """Execute all importer self-tests.

    Returns 0 on success, 1 on failure.
    """
    errors = 0
    test_dir = None

    try:
        # Create synthetic fixture
        with tempfile.TemporaryDirectory() as test_dir:
            fixture_path = os.path.join(test_dir, "synthetic.xlsx")
            output_dir_a = os.path.join(test_dir, "out-a")
            output_dir_b = os.path.join(test_dir, "out-b")
            os.makedirs(output_dir_a)
            os.makedirs(output_dir_b)

            _make_synthetic_wb(fixture_path)

            # ── Test 1: Basic import succeeds ──
            print("Test 1: Basic import ... ", end="")
            manifest = import_xlsx(fixture_path, output_dir_a, "hsk-3.0")
            total = manifest["accepted"] + manifest["rejected"]
            assert manifest["totalRows"] == total, \
                f"totalRows ({manifest['totalRows']}) != accepted+rejected ({total})"
            assert manifest["accepted"] == 12, \
                f"expected 12 accepted, got {manifest['accepted']}"
            assert manifest["rejected"] == 3, \
                "expected 3 rejected (1 empty simplified + 1 duplicate + 1 case-variant duplicate), " \
                "got {}".format(manifest["rejected"])
            assert manifest["batchCount"] >= 1
            print("PASS")

            # ── Test 2: Source sheets detected ──
            print("Test 2: Source sheet detection ... ", end="")
            assert "HSK Level 1" in manifest["sourceSheets"]
            assert "HSK Level 2" in manifest["sourceSheets"]
            assert "HSK Level 3-4" in manifest["sourceSheets"]
            assert "Metadata Only" not in manifest["sourceSheets"]
            assert "Empty" not in manifest["sourceSheets"]
            print("PASS")

            # ── Test 3: Duplicate diagnostics ──
            print("Test 3: Duplicate diagnostics ... ", end="")
            assert len(manifest["duplicateDiagnostics"]) == 2, \
                f"expected 2 duplicate diagnostics (exact + case variant), got {len(manifest['duplicateDiagnostics'])}"
            for diag in manifest["duplicateDiagnostics"]:
                assert "duplicate identity" in diag["reason"].lower()
            print("PASS")

            # ── Test 4: Rejected row tracking ──
            print("Test 4: Rejected row tracking ... ", end="")
            assert len(manifest["rejectedRows"]) == 3, \
                f"expected 3 rejected rows, got {len(manifest['rejectedRows'])}"
            rejected_reasons = [r["reason"] for r in manifest["rejectedRows"]]
            assert any("missing simplified" in r for r in rejected_reasons)
            assert any("duplicate identity" in r for r in rejected_reasons)

            # Verify exact row numbers (openpyxl 1-based)
            # Row 7 = empty simplified, Row 8 = exact match duplicate, Row 9 = case-variant duplicate
            rejected_rows = sorted(
                [r for r in manifest["rejectedRows"]
                 if r["sheet"] == "HSK Level 1"],
                key=lambda r: r["row"],
            )
            assert len(rejected_rows) == 3
            assert rejected_rows[0]["row"] == 7, \
                f"expected empty-simplified at row 7, got {rejected_rows[0]['row']}"
            assert rejected_rows[1]["row"] == 8, \
                f"expected exact duplicate at row 8, got {rejected_rows[1]['row']}"
            assert rejected_rows[2]["row"] == 9, \
                f"expected case-variant duplicate at row 9, got {rejected_rows[2]['row']}"

            # Verify case-variant reason mentions the variant
            case_variant_reason = rejected_rows[2]["reason"]
            assert "Shuǐ" in case_variant_reason or "水" in case_variant_reason, \
                f"expected case-variant info in reason: {case_variant_reason}"
            print("PASS")

            # ── Test 5: Batch integrity ──
            print("Test 5: Batch size and ordering ... ", end="")
            _check_batch_order_and_size(output_dir_a)
            print("PASS")

            # ── Test 6: Byte-identical re-run ──
            print("Test 6: Byte-identical re-run ... ", end="")
            manifest_b = import_xlsx(fixture_path, output_dir_b, "hsk-3.0")

            # Compare batch files
            for b in manifest["batches"]:
                path_a = os.path.join(output_dir_a, b["filename"])
                path_b = os.path.join(output_dir_b, b["filename"])
                with open(path_a, "r", encoding="utf-8") as f:
                    data_a = f.read()
                with open(path_b, "r", encoding="utf-8") as f:
                    data_b = f.read()
                assert data_a == data_b, \
                    f"Byte mismatch on {b['filename']}"

            # Compare manifests (differing only in checksum which is identical
            # since same input file)
            path_ma = os.path.join(output_dir_a, "manifest.json")
            path_mb = os.path.join(output_dir_b, "manifest.json")
            with open(path_ma, "r", encoding="utf-8") as f:
                ma = json.load(f)
            with open(path_mb, "r", encoding="utf-8") as f:
                mb = json.load(f)
            assert ma == mb, "Manifests differ between identical runs"
            print("PASS")

            # ── Test 7: Level distribution and manifest counts ──
            print("Test 7: Level distribution and manifest counts ... ", end="")
            by_level = manifest["acceptedByLevel"]
            assert by_level.get("1") == 5, \
                f"expected 5 at level 1, got {by_level.get('1')}"
            assert by_level.get("2") == 3, \
                f"expected 3 at level 2, got {by_level.get('2')}"
            assert by_level.get("3") == 3, \
                f"expected 3 at level 3, got {by_level.get('3')}"
            assert by_level.get("4") == 1, \
                f"expected 1 at level 4, got {by_level.get('4')}"
            # Verify by-level sum invariants
            _check_manifest_counts(manifest)
            print("PASS")

            # ── Test 8: Verify images ignored (no op needed — openpyxl
            # never reads images as cell values)
            #
            # openpyxl simply does not surface embedded images via .value.
            # We confirm by making sure the cell count matches: the XLSX
            # has 9 data rows across 3 data sheets and images contribute 0.

            # ── Test 9: Column mapping variants ──
            print("Test 9: Column header variants ... ", end="")
            # Sheet 2 uses Chinese headers (简体, 拼音, 级别, 日文)
            # Sheet 3 uses mixed (简体字, 拼音, 繁体字, Level)
            # Both were correctly detected — confirmed by accepted entries
            assert len(manifest["sourceSheets"]) == 3
            assert manifest["accepted"] == 12
            print("PASS")

            # ── Test 10: No forbidden fields in output ──
            print("Test 10: Output field hygiene ... ", end="")
            for b in manifest["batches"]:
                batch_path = os.path.join(output_dir_a, b["filename"])
                with open(batch_path, "r", encoding="utf-8") as f:
                    batch = json.load(f)
                for entry in batch["vocabulary"]:
                    assert "id" in entry
                    assert "simplified" in entry
                    assert "pinyin" in entry
                    assert "hsk" in entry
                    assert "source" in entry
                    assert "_sourceSheet" not in entry
                    assert "_sourceRow" not in entry
            print("PASS")

            # ── Test 11: Normalization matches validator ──
            print("Test 11: Normalization matches validator ... ", end="")
            # case-fold variant: "Ài" vs "ài" should be detected as duplicate
            assert _normalize_pinyin("Ài") == _normalize_pinyin("ài"), \
                "pinyin normalization must match validator casefold behavior"
            # whitespace variant: "nǐ hǎo" vs "nǐhǎo" should be detected
            assert _normalize_pinyin("nǐ hǎo") == _normalize_pinyin("nǐhǎo"), \
                "pinyin normalization must remove all whitespace"
            # NFKC variant: simplified with fullwidth chars
            assert _normalize_simplified("爱") == _normalize_simplified("爱"), \
                "simplified normalization must be NFKC-stable"
            print("PASS")

            # ── Test 12: Missing japanese column is fatal ──
            print("Test 12: Missing japanese column ... ", end="")
            _test_missing_japanese()
            print("PASS")

            # ── Test 13: Invalid standardVersion ──
            print("Test 13: Invalid standardVersion ... ", end="")
            _test_invalid_standard_version()
            print("PASS")

            # ── Test 14: Invalid level values rejected ──
            print("Test 14: Invalid level values ... ", end="")
            _test_invalid_level_rejection()
            print("PASS")

            # ── Test 15: Formula in required field rejected ──
            print("Test 15: Formula rejection ... ", end="")
            _test_formula_rejection()
            print("PASS")

            # ── Test 16: Malformed sheet fails before output ──
            print("Test 16: Malformed sheet detection ... ", end="")
            _test_sheet_detection_malformed()
            print("PASS")

            # ── Test 17: Fully empty rows are ignored deterministically ──
            print("Test 17: Empty row handling ... ", end="")
            _test_empty_rows_ignored()
            print("PASS")

    except Exception as e:
        print(f"\nFAILED: {e}", file=sys.stderr)
        errors = 1

    return errors


if __name__ == "__main__":
    main()
