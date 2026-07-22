#!/usr/bin/env python3
"""
Teacher vocabulary XLSX Importer for Chabiko (#112).

Reads the teacher-curriculum workbook `单词表(带图).xlsx`, parses text-only
content from exact sheet/column mappings, and emits deterministic draft
teacher-curriculum vocabulary records plus a bounded import manifest.

Usage:
    uv run python scripts/import-teacher-vocabulary-xlsx.py <input.xlsx> <output-dir>
    uv run python scripts/import-teacher-vocabulary-xlsx.py --test
"""

import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import unicodedata

import openpyxl
from openpyxl import Workbook


# ─── Exact sheet classification ──────────────────────────────────────────────

SHEET_MAP = {
    "名词1":     {"difficultyBand": "star-1", "sourceDifficultyLabel": "☆", "partOfSpeech": "noun"},
    "动词1":     {"difficultyBand": "star-1", "sourceDifficultyLabel": "☆", "partOfSpeech": "verb"},
    "形容词1":   {"difficultyBand": "star-1", "sourceDifficultyLabel": "☆", "partOfSpeech": "adjective"},
    "副词":      {"difficultyBand": "star-1", "sourceDifficultyLabel": "☆", "partOfSpeech": "adverb"},
    "名词2":     {"difficultyBand": "star-2", "sourceDifficultyLabel": "☆☆", "partOfSpeech": "noun"},
    "形容词2":   {"difficultyBand": "star-2", "sourceDifficultyLabel": "☆☆", "partOfSpeech": "adjective"},
    "动词2":     {"difficultyBand": "star-2", "sourceDifficultyLabel": "☆☆", "partOfSpeech": "verb"},
}

SHEET_ORDER = ["名词1", "动词1", "形容词1", "副词", "名词2", "形容词2", "动词2"]

SEPARATOR_SHEETS = frozenset({"難易度☆", "難易度☆☆"})

POS_ORDER = {"noun": 0, "verb": 1, "adjective": 2, "adverb": 3}
DIFFICULTY_ORDER = {"star-1": 0, "star-2": 1}

# ─── Exact header mappings ───────────────────────────────────────────────────

REQUIRED_HEADERS = {"单词", "拼音", "日语翻译", "难易度"}
IGNORED_HEADERS = {"造词/造句", "日文字", "备注"}
ALL_KNOWN_HEADERS = REQUIRED_HEADERS | IGNORED_HEADERS

HEADER_TO_CANONICAL = {
    "单词": "simplified",
    "拼音": "pinyin",
    "日语翻译": "japanese",
    "难易度": "difficulty_check",
}


# ─── Unicode/whitespace normalization ────────────────────────────────────────

def normalize_text(value):
    """Normalize whitespace and Unicode deterministically."""
    if not isinstance(value, str):
        value = str(value) if value is not None else ""
    value = unicodedata.normalize("NFC", value.strip())
    value = re.sub(r"\s+", " ", value)
    return value


def _normalize_simplified(text):
    """Normalize simplified Chinese per teacher identity contract.

    Matches validate-content-schema.py _normalize_simplified exactly.
    """
    normalized = unicodedata.normalize("NFKC", text)
    return "".join(ch for ch in normalized if not ch.isspace())


def _normalize_pinyin(text):
    """Normalize pinyin per teacher identity contract.

    Matches validate-content-schema.py _normalize_pinyin exactly.
    """
    normalized = unicodedata.normalize("NFKC", text)
    case_folded = normalized.casefold()
    return "".join(ch for ch in case_folded if not ch.isspace())


# ─── Stable ID generation ────────────────────────────────────────────────────

def generate_stable_id(simplified, pinyin, difficulty_band):
    """Generate a deterministic vocabulary ID.

    Algorithm: SHA-256 of "teacher-core-v1|{normalizedSimplified}|{normalizedPinyin}"
    → first 12 hex chars.
    """
    norm_s = _normalize_simplified(simplified)
    norm_p = _normalize_pinyin(pinyin)
    seed = f"teacher-core-v1|{norm_s}|{norm_p}"
    hash_digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    return f"teacher-{difficulty_band}-{hash_digest}"


def generate_illustration_id(vocabulary_id):
    """Generate a deterministic illustration ID from a vocabulary ID."""
    return f"ill-{vocabulary_id}"


# ─── Sheet classification ────────────────────────────────────────────────────

def classify_sheets(wb):
    """Classify workbook sheets into data sheets, separator sheets, and unknowns.

    Returns (data_sheets, separator_sheets).

    Unknown non-empty sheets raise ValueError before any output mutation.
    """
    data_sheets = []
    separator_sheets = []

    for ws in wb.worksheets:
        name = ws.title

        if name in SEPARATOR_SHEETS:
            separator_sheets.append(name)
            continue

        classification = SHEET_MAP.get(name)
        if classification is not None:
            data_sheets.append((name, ws, classification))
            continue

        if _sheet_has_any_data(ws):
            raise ValueError(
                f"Unknown non-empty sheet(s) found: '{name}'. "
                f"Expected one of: {', '.join(sorted(SHEET_MAP.keys()))}"
            )

    return data_sheets, separator_sheets


def _sheet_has_any_data(ws):
    """Check if a worksheet has any cell content."""
    if ws.max_row is None or ws.max_column is None:
        return False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                return True
    return False


# ─── Header resolution ──────────────────────────────────────────────────────

def resolve_teacher_headers(header_row, sheet_name=""):
    """Parse headers from a data sheet header row.

    Returns (canonical_mapping, ignored_column_names, unknown_column_names).

    Raises ValueError if a required header appears more than once.
    """
    canonical_mapping = {}
    ignored_headers_found = []
    unknown_headers = []

    for idx, raw in enumerate(header_row):
        text = normalize_text(raw).strip() if raw else ""
        if not text:
            continue

        if text in REQUIRED_HEADERS:
            canonical = HEADER_TO_CANONICAL[text]
            if canonical in canonical_mapping:
                prev_pos = canonical_mapping[canonical]
                loc = f"'{sheet_name}'" if sheet_name else "header row"
                raise ValueError(
                    f"Duplicate required header '{text}' (mapped to '{canonical}') "
                    f"at columns {prev_pos + 1} and {idx + 1} in sheet {loc}"
                )
            canonical_mapping[canonical] = idx
        elif text in IGNORED_HEADERS:
            ignored_headers_found.append(text)
        else:
            unknown_headers.append(text)

    return canonical_mapping, ignored_headers_found, unknown_headers


def check_required_headers(mapping):
    """Return list of missing required canonical keys."""
    required_keys = {"simplified", "pinyin", "japanese", "difficulty_check"}
    return [k for k in required_keys if k not in mapping]


# ─── Row parsing ─────────────────────────────────────────────────────────────

def parse_teacher_row(ws, row_idx, col_mapping):
    """Parse a single data row.

    Returns (record_dict, has_formula_required).
    """
    record = {"_sourceSheet": ws.title, "_sourceRow": row_idx}

    for canonical_key, col_idx in col_mapping.items():
        cell = ws.cell(row=row_idx, column=col_idx + 1)
        raw = cell.value
        is_formula = (cell.data_type == "f")

        if is_formula:
            record[canonical_key] = ""
            if canonical_key in ("simplified", "pinyin", "japanese", "difficulty_check"):
                record.setdefault("_formula_fields", []).append(canonical_key)
        elif raw is None:
            record[canonical_key] = ""
        elif isinstance(raw, bool):
            record[canonical_key] = str(raw)
        else:
            record[canonical_key] = normalize_text(str(raw))

    return record, "_formula_fields" in record


def _is_row_fully_empty(ws, row_idx, max_col):
    """Check if every cell in the row is None."""
    for col in range(1, max_col + 1):
        if ws.cell(row=row_idx, column=col).value is not None:
            return False
    return True


# ─── Record validation ───────────────────────────────────────────────────────

def validate_teacher_record(record, classification, seen_ids, seen_identities):
    """Validate and convert a parsed record into a vocabulary dict.

    Returns (vocab_dict, rejection_str) or (None, rejection_str).
    """
    simplified = record.get("simplified", "").strip()
    pinyin = record.get("pinyin", "").strip()
    japanese = record.get("japanese", "").strip()
    diff_check = record.get("difficulty_check", "").strip()
    sheet = record.get("_sourceSheet", "")
    row = record.get("_sourceRow", 0)
    formula_fields = record.get("_formula_fields", [])

    if formula_fields:
        fields_str = ", ".join(formula_fields)
        return None, f"formula in required column(s) {fields_str}: {sheet}:{row}"

    if not simplified:
        return None, f"missing simplified: {sheet}:{row}"
    if not pinyin:
        return None, f"missing pinyin: {sheet}:{row}"
    if not japanese:
        return None, f"missing japanese: {sheet}:{row}"
    if not diff_check:
        return None, f"missing difficulty check: {sheet}:{row}"

    expected_label = classification["sourceDifficultyLabel"]
    norm_diff = normalize_text(diff_check)
    if norm_diff != expected_label:
        return None, (
            f"difficulty mismatch: expected '{expected_label}', "
            f"got '{norm_diff}': {sheet}:{row}"
        )

    identity = (_normalize_simplified(simplified), _normalize_pinyin(pinyin))
    if identity in seen_identities:
        return None, (
            f"duplicate identity ({simplified}, {pinyin}): "
            f"{sheet}:{row}"
        )

    difficulty_band = classification["difficultyBand"]
    vid = generate_stable_id(simplified, pinyin, difficulty_band)
    if vid in seen_ids:
        return None, f"collision on ID {vid}: {sheet}:{row}"

    vocab = {
        "id": vid,
        "simplified": simplified,
        "simplifiedStatus": "authored",
        "pinyin": pinyin,
        "japanese": japanese,
        "source": {"type": "teacher-workbook"},
        "reviewStatus": "draft",
        "curriculum": {
            "sourceId": "teacher-core-v1",
            "difficultyBand": difficulty_band,
            "sourceDifficultyLabel": classification["sourceDifficultyLabel"],
            "partOfSpeech": classification["partOfSpeech"],
            "sourceSheet": sheet,
            "sourceRow": row,
        },
    }

    return vocab, None


# ─── Batch output ────────────────────────────────────────────────────────────

def write_teacher_batches(accepted_entries, output_dir):
    """Write vocabulary records to deterministic batch JSON files.

    *accepted_entries* is a list of ``(vocab_dict, source_sheet, source_row)``
    tuples.  Returns batch metadata list for the manifest.
    """
    accepted_entries.sort(key=_teacher_sort_key)

    batches = []
    for i in range(0, len(accepted_entries), 50):
        chunk = accepted_entries[i:i + 50]
        batch_num = i // 50 + 1
        filename = f"teacher-vocabulary-batch-{batch_num:02d}.json"
        filepath = os.path.join(output_dir, filename)

        batch_vocab = []
        batch_sources = []
        for v, sheet, row in chunk:
            batch_vocab.append(v)
            batch_sources.append({
                "id": v["id"],
                "sheet": sheet,
                "row": row,
            })

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({"vocabulary": batch_vocab}, f,
                      ensure_ascii=False, indent=2)
            f.write("\n")

        batches.append({
            "filename": filename,
            "batchNumber": batch_num,
            "entryCount": len(batch_vocab),
            "firstEntryId": batch_vocab[0]["id"],
            "lastEntryId": batch_vocab[-1]["id"],
            "sourceRows": batch_sources,
        })

    return batches


def _teacher_sort_key(entry):
    """Stable sort key for accepted entries.

    1. difficulty: star-1, then star-2
    2. part of speech: noun, verb, adjective, adverb
    3. source sheet in fixed table order
    4. numeric source row ascending
    """
    v, sheet, row = entry
    cur = v["curriculum"]
    diff_order = DIFFICULTY_ORDER.get(cur["difficultyBand"], 99)
    pos_order = POS_ORDER.get(cur["partOfSpeech"], 99)
    sheet_order = SHEET_ORDER.index(sheet) if sheet in SHEET_ORDER else 99
    return (diff_order, pos_order, sheet_order, row)


# ─── Manifest helpers ─────────────────────────────────────────────────────────

def _count_by_dimension(entries, key_fn):
    """Count entries by a dimension key extracted via key_fn."""
    counts = {}
    for entry in entries:
        k = key_fn(entry)
        counts[k] = counts.get(k, 0) + 1
    return counts


def _dimension_sort_key(item):
    """Sort dimension keys: difficulties first, then POS."""
    k, _ = item
    parts = k.split(":", 1)
    if len(parts) == 2:
        dk = DIFFICULTY_ORDER.get(parts[0], 99)
        pk = POS_ORDER.get(parts[1], 99)
        return (dk, pk)
    return (99, 99)


def _serialise_dimension_counts(counter):
    """Serialise difficulty:pos count dict with deterministic order."""
    items = sorted(counter.items(), key=_dimension_sort_key)
    return {k: v for k, v in items}


def _dimension_key(classification):
    """Create a 'difficulty:pos' key from a classification dict."""
    return f"{classification['difficultyBand']}:{classification['partOfSpeech']}"


# ─── Output directory safety ─────────────────────────────────────────────────

def prepare_output_dir(output_dir):
    """Prepare the output directory with fail-fast semantics.

    - Does not exist → creates it.
    - Exists and is empty (no files nor directories inside) → allowed.
    - Exists and is non-empty, or exists but is not a directory → raises ValueError
      before any output mutation.
    - Never deletes user files.
    """
    if os.path.exists(output_dir):
        if not os.path.isdir(output_dir):
            raise ValueError(
                f"Output path '{output_dir}' exists but is not a directory"
            )
        entries = os.listdir(output_dir)
        if entries:
            raise ValueError(
                f"Output directory '{output_dir}' already exists and is "
                f"non-empty ({len(entries)} entries). Refusing to overwrite."
            )
    else:
        os.makedirs(output_dir)


# ─── Main import pipeline ────────────────────────────────────────────────────

def import_xlsx(input_path, output_dir):
    """Run the full teacher-workbook import pipeline.

    Returns the manifest dict.
    """
    # 1. Source checksum (before any output mutation)
    with open(input_path, "rb") as f:
        checksum = hashlib.sha256(f.read()).hexdigest()

    # 2. Open workbook (disable data_only to detect formula cells)
    wb = openpyxl.load_workbook(input_path, data_only=False)

    # 3. Classify sheets (raises ValueError for unknown non-empty sheets)
    data_sheets, separator_sheets = classify_sheets(wb)
    if not data_sheets:
        raise ValueError("No data sheets found")

    # 4. Parse headers for each data sheet
    sheet_info = []
    for name, ws, classification in data_sheets:
        header_row = [cell.value for cell in ws[1]]
        col_mapping, ignored_cols, unknown_cols = resolve_teacher_headers(header_row, sheet_name=name)

        missing = check_required_headers(col_mapping)
        if missing:
            raise ValueError(
                f"Sheet '{name}' is missing required header(s): "
                f"{', '.join(missing)}"
            )

        sheet_info.append((name, ws, classification, col_mapping, ignored_cols, unknown_cols))

    # Build ignored-columns-by-sheet for manifest
    ignored_columns_by_sheet = {}
    for name, _ws, _class, _cm, ignored_cols, unknown_cols in sheet_info:
        all_ignored = list(ignored_cols) + list(unknown_cols)
        if all_ignored:
            ignored_columns_by_sheet[name] = all_ignored

    detected_separator_sheets = list(separator_sheets) if separator_sheets else []

    # 5. Parse all rows and validate
    seen_ids = set()
    seen_identities = set()
    accepted_entries = []
    rejected = []
    total_candidate = 0

    for name, ws, classification, col_mapping, _ig, _uk in sheet_info:
        max_col = ws.max_column or 1
        for row_idx in range(2, ws.max_row + 1):
            if _is_row_fully_empty(ws, row_idx, max_col):
                continue

            total_candidate += 1

            record, _ = parse_teacher_row(ws, row_idx, col_mapping)

            vocab, reason = validate_teacher_record(
                record, classification, seen_ids, seen_identities,
            )

            if reason:
                rejected.append({
                    "reason": reason,
                    "sheet": name,
                    "row": row_idx,
                })
            else:
                normalized_identity = (
                    _normalize_simplified(vocab["simplified"]),
                    _normalize_pinyin(vocab["pinyin"]),
                )
                seen_ids.add(vocab["id"])
                seen_identities.add(normalized_identity)
                accepted_entries.append((vocab, name, row_idx))

    # Count by dimension
    total_by_dim = {}
    for name, ws, classification, col_mapping, _ig, _uk in sheet_info:
        key = _dimension_key(SHEET_MAP[name])
        if key not in total_by_dim:
            total_by_dim[key] = 0
        max_col = ws.max_column or 1
        for row_idx in range(2, ws.max_row + 1):
            if not _is_row_fully_empty(ws, row_idx, max_col):
                total_by_dim[key] = total_by_dim.get(key, 0) + 1

    accepted_by_dim = _count_by_dimension(
        accepted_entries,
        lambda x: _dimension_key(SHEET_MAP.get(x[1], {})),
    )

    rejected_by_dim = {}
    for r in rejected:
        sheet = r["sheet"]
        if sheet in SHEET_MAP:
            key = _dimension_key(SHEET_MAP[sheet])
            rejected_by_dim[key] = rejected_by_dim.get(key, 0) + 1

    # 6. Validate output directory (before any output mutation)
    prepare_output_dir(output_dir)

    # 7. Write batches (output mutation starts here)"
    batches = write_teacher_batches(accepted_entries, output_dir)

    # 7. Build manifest
    duplicate_diagnostics = [
        r for r in rejected if "duplicate" in r["reason"]
    ]

    accepted_items = [
        {
            "vocabularyId": v["id"],
            "expectedIllustrationId": generate_illustration_id(v["id"]),
            "sourceSheet": sheet,
            "sourceRow": row,
        }
        for v, sheet, row in accepted_entries
    ]

    manifest = {
        "sourceFile": os.path.basename(input_path),
        "sourceChecksumSha256": checksum,
        "sourceId": "teacher-core-v1",
        "detectedSheets": [s[0] for s in sheet_info],
        "separatorSheets": detected_separator_sheets,
        "ignoredColumnsBySheet": ignored_columns_by_sheet,
        "totalRows": total_candidate,
        "totalByDifficultyAndPartOfSpeech": _serialise_dimension_counts(total_by_dim),
        "accepted": len(accepted_entries),
        "acceptedByDifficultyAndPartOfSpeech": _serialise_dimension_counts(accepted_by_dim),
        "rejected": len(rejected),
        "rejectedByDifficultyAndPartOfSpeech": _serialise_dimension_counts(rejected_by_dim),
        "duplicateDiagnostics": duplicate_diagnostics,
        "rejectedRows": rejected,
        "batchCount": len(batches),
        "batches": batches,
        "acceptedItems": accepted_items,
    }

    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
        f.write("\n")

    # 8. Validate every batch against the #74 contract
    _validate_batches(output_dir, batches)

    return manifest


def _validate_batches(output_dir, batches):
    """Run the #74 content schema validator on every batch file."""
    validator = os.path.join(os.path.dirname(__file__), "validate-content-schema.py")
    for b in batches:
        batch_path = os.path.join(output_dir, b["filename"])
        result = subprocess.run(
            [sys.executable, validator, "--check", batch_path],
            capture_output=True, text=True,
        )
        if result.returncode != 0:
            lines = [l for l in result.stdout.strip().split("\n") if l.strip()]
            summary = "; ".join(lines[:3])
            raise RuntimeError(
                f"Batch {b['filename']} failed #74 contract validation: {summary}"
            )


# ─── CLI entry point ─────────────────────────────────────────────────────────

def main():
    if len(sys.argv) == 2 and sys.argv[1] == "--test":
        sys.exit(run_tests())

    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.isfile(input_path):
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        manifest = import_xlsx(input_path, output_dir)
        print(f"Import complete: {manifest['accepted']} accepted, "
              f"{manifest['rejected']} rejected")
        print(f"  Batches: {manifest['batchCount']}")
        print(f"  Output: {output_dir}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# Self-tests
# ═══════════════════════════════════════════════════════════════════════════════

def _make_synthetic_wb(path):
    """Create a synthetic multi-sheet XLSX fixture at *path*."""
    wb = Workbook()

    default = wb.active
    wb.remove(default)

    # Data sheet: 名词1
    ws1 = wb.create_sheet("名词1")
    ws1.append(["单词", "拼音", "日语翻译", "难易度", "造词/造句", "日文字", "备注"])
    ws1.append(["爱", "ài", "愛する", "☆", "愛情", "あい", ""])
    ws1.append(["猫", "māo", "猫", "☆", "", "ねこ", "動物"])
    ws1.append(["狗", "gǒu", "犬", "☆", "", "", "ペット"])

    # Row with empty simplified
    ws1.append(["", "shénme", "何", "☆", "", "", ""])
    # Duplicate identity (爱 + ài)
    ws1.append(["爱", "ài", "愛情", "☆", "", "", ""])
    # Case-variant duplicate (狗 + Gǒu)
    ws1.append(["狗", "Gǒu", "イヌ", "☆", "", "", ""])
    # Fully empty row
    ws1.append([None, None, None, None, None, None, None])
    # Formula in required column
    ws1.append(["test", "cèshì", "テスト", "☆", "", "", ""])

    # Data sheet: 动词1
    ws2 = wb.create_sheet("动词1")
    ws2.append(["单词", "拼音", "日语翻译", "难易度"])
    ws2.append(["吃", "chī", "食べる", "☆"])
    ws2.append(["喝", "hē", "飲む", "☆"])

    # Separator sheet
    ws3 = wb.create_sheet("難易度☆☆")
    ws3.append(["separator data"])

    # Data sheet: 名词2
    ws4 = wb.create_sheet("名词2")
    ws4.append(["单词", "拼音", "日语翻译", "难易度"])
    ws4.append(["学校", "xuéxiào", "学校", "☆☆"])

    return wb


def _inject_formula(path, sheet_name, row, col, formula):
    """Inject a formula into an existing XLSX file."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    ws.cell(row=row, column=col).value = formula
    wb.save(path)


def _check_batch_order_and_size(output_dir):
    """Verify every batch contains 1-50 entries with stable ordering."""
    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)

    assert manifest["batchCount"] >= 1, "must produce at least one batch"

    for b in manifest["batches"]:
        assert 1 <= b["entryCount"] <= 50, \
            f"batch {b['batchNumber']} has {b['entryCount']} entries (want 1-50)"

        batch_path = os.path.join(output_dir, b["filename"])
        with open(batch_path, "r", encoding="utf-8") as f:
            batch = json.load(f)

        assert len(batch["vocabulary"]) == b["entryCount"]

        entries = batch["vocabulary"]
        for i in range(1, len(entries)):
            prev = entries[i - 1]
            cur = entries[i]
            prev_cur = prev["curriculum"]
            cur_cur = cur["curriculum"]
            prev_key = (
                DIFFICULTY_ORDER.get(prev_cur["difficultyBand"], 99),
                POS_ORDER.get(prev_cur["partOfSpeech"], 99),
                SHEET_ORDER.index(prev_cur["sourceSheet"]) if prev_cur["sourceSheet"] in SHEET_ORDER else 99,
                prev_cur["sourceRow"],
            )
            cur_key = (
                DIFFICULTY_ORDER.get(cur_cur["difficultyBand"], 99),
                POS_ORDER.get(cur_cur["partOfSpeech"], 99),
                SHEET_ORDER.index(cur_cur["sourceSheet"]) if cur_cur["sourceSheet"] in SHEET_ORDER else 99,
                cur_cur["sourceRow"],
            )
            assert prev_key <= cur_key, \
                f"batch {b['batchNumber']} not sorted at index {i}: {prev_key} > {cur_key}"

    return manifest


def _check_manifest_counts(manifest):
    """Verify by-dimension count sums equal their top-level totals."""
    total_dim = sum(manifest["totalByDifficultyAndPartOfSpeech"].values())
    assert total_dim == manifest["totalRows"], \
        f"totalByDimension sum ({total_dim}) != totalRows ({manifest['totalRows']})"

    accepted_dim = sum(manifest["acceptedByDifficultyAndPartOfSpeech"].values())
    assert accepted_dim == manifest["accepted"], \
        f"acceptedByDimension sum ({accepted_dim}) != accepted ({manifest['accepted']})"

    rejected_dim = sum(manifest["rejectedByDifficultyAndPartOfSpeech"].values())
    assert rejected_dim == manifest["rejected"], \
        f"rejectedByDimension sum ({rejected_dim}) != rejected ({manifest['rejected']})"


# ─── Phase 1 tests: Workbook parsing ─────────────────────────────────────────

def _test_sheet_classification():
    """Verify exact sheet classification, separator handling, and unknown fatal."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _default = _wb.active
        _wb.remove(_default)
        _ws_noun = _wb.create_sheet("名词1")
        _ws_noun.append(["单词", "拼音", "日语翻译", "难易度", "a"])
        _ws_noun.append(["爱", "ài", "愛する", "☆", ""])
        _ws2 = _wb.create_sheet("難易度☆")
        _ws2.append(["x"])
        _ws3 = _wb.create_sheet("副词")
        _ws3.append(["单词", "拼音", "日语翻译", "难易度", "a"])
        _ws3.append(["一直", "yīzhí", "いつも", "☆", ""])
        _path = os.path.join(_td, "good.xlsx")
        _wb.save(_path)
        _wb2 = openpyxl.load_workbook(_path)
        _data, _sep = classify_sheets(_wb2)
        _names = sorted(s[0] for s in _data)
        assert _names == sorted(["名词1", "副词"]), f"Expected ['名词1', '副词'], got {_names}"
        assert _sep == ["難易度☆"], f"Expected ['難易度☆'], got {_sep}"

        # Unknown non-empty sheet → fatal
        _wb3 = Workbook()
        _default2 = _wb3.active
        _wb3.remove(_default2)
        _ws4 = _wb3.create_sheet("名词1")
        _ws4.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws4.append(["爱", "ài", "愛する", "☆"])
        _ws5 = _wb3.create_sheet("Unknown")
        _ws5.append(["data"])
        _path2 = os.path.join(_td, "unknown.xlsx")
        _wb3.save(_path2)
        _wb4 = openpyxl.load_workbook(_path2)
        try:
            classify_sheets(_wb4)
            raise AssertionError("expected ValueError for unknown sheet")
        except ValueError as e:
            assert "Unknown" in str(e)

        # Empty sheet → silently ignored
        _wb5 = Workbook()
        _default3 = _wb5.active
        _wb5.remove(_default3)
        _ws6 = _wb5.create_sheet("名词1")
        _ws6.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws6.append(["爱", "ài", "愛する", "☆"])
        _empty_sheet = _wb5.create_sheet("EmptySheet")
        _ = _empty_sheet  # no data → empty
        _wb5.save(os.path.join(_td, "empty_sheet.xlsx"))
        _wb6 = openpyxl.load_workbook(os.path.join(_td, "empty_sheet.xlsx"))
        _data2, _ = classify_sheets(_wb6)
        assert len(_data2) == 1, f"Expected 1 data sheet, got {len(_data2)}"


def _test_exact_headers():
    """Verify exact header detection works."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _header = [cell.value for cell in _ws[1]]
        _mapping, _ignored, _unknown = resolve_teacher_headers(_header)
        assert _mapping.get("simplified") == 0
        assert _mapping.get("pinyin") == 1
        assert _mapping.get("japanese") == 2
        assert _mapping.get("difficulty_check") == 3
        assert _ignored == []

        # Missing header
        _wb2 = Workbook()
        _ws2 = _wb2.active
        _ws2.title = "名词1"
        _ws2.append(["单词", "拼音", "难易度"])
        _header2 = [cell.value for cell in _ws2[1]]
        _mapping2, _, _ = resolve_teacher_headers(_header2)
        _missing = check_required_headers(_mapping2)
        assert "japanese" in _missing

        # Unknown extra headers
        _wb3 = Workbook()
        _ws3 = _wb3.active
        _ws3.title = "名词1"
        _ws3.append(["单词", "拼音", "日语翻译", "难易度", "未知列"])
        _header3 = [cell.value for cell in _ws3[1]]
        _mapping3, _ignored3, _unknown3 = resolve_teacher_headers(_header3)
        assert "未知列" in _unknown3


def _test_ignored_headers():
    """Verify ignored headers (造词/造句, 日文字, 备注) are tracked."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度", "造词/造句", "日文字", "备注"])
        _header = [cell.value for cell in _ws[1]]
        _mapping, _ignored, _unknown = resolve_teacher_headers(_header)
        assert "造词/造句" in _ignored
        assert "日文字" in _ignored
        assert "备注" in _ignored
        assert _unknown == []


def _test_empty_partial_rows():
    """Verify fully empty rows are ignored, partial rows are tracked."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "rows.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _ws.append(["", "", "", ""])
        _ws.append([None, None, None, None])
        _ws.append(["猫", "māo", "", "☆"])
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        # Row with ['','','',''] saved as all-None by openpyxl → empty
        # So only row 2 (valid) and row 5 (missing japanese) are candidates
        assert _m["totalRows"] == 2
        assert _m["accepted"] == 1
        assert _m["rejected"] == 1


def _test_required_formula_rejection():
    """Verify formulas in required columns reject the row."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "formula.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _wb.save(_path)
        _inject_formula(_path, "名词1", 2, 1, "=A2&B2")
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["accepted"] == 0
        assert _m["rejected"] == 1
        assert "formula" in _m["rejectedRows"][0]["reason"].lower()


def _test_unknown_column_reporting():
    """Verify unknown extra columns are listed in ignoredColumnsBySheet."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度", "ExtraCol", "Another"])
        _ws.append(["爱", "ài", "愛する", "☆", "", ""])
        _path = os.path.join(_td, "unknown_col.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert "名词1" in _m["ignoredColumnsBySheet"]
        _cols = _m["ignoredColumnsBySheet"]["名词1"]
        assert "ExtraCol" in _cols
        assert "Another" in _cols


def _test_unknown_sheet_fatal():
    """Verify unknown non-empty sheet raises ValueError before output."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _default = _wb.active
        _wb.remove(_default)
        _ws1 = _wb.create_sheet("名词1")
        _ws1.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws1.append(["爱", "ài", "愛する", "☆"])
        _wb.create_sheet("UnknownSheet").append(["data"])
        _path = os.path.join(_td, "unknown_sheet.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        try:
            import_xlsx(_path, _out)
            raise AssertionError("expected ValueError for unknown sheet")
        except ValueError:
            assert not os.path.exists(os.path.join(_out, "manifest.json")), \
                "output created before sheet validation"


# ─── Phase 2 tests: Records and deterministic identity ──────────────────────

def _test_normalization():
    """Verify NFKC + Unicode-whitespace normalization matches validator."""
    assert _normalize_simplified("爱") == _normalize_simplified("爱")
    assert _normalize_simplified("你 好") == _normalize_simplified("你好")
    assert _normalize_pinyin("Ài") == _normalize_pinyin("ài")
    assert _normalize_pinyin("nǐ hǎo") == _normalize_pinyin("nǐhǎo")


def _test_vocabulary_id():
    """Verify vocabulary ID generation algorithm."""
    _id1 = generate_stable_id("爱", "ài", "star-1")
    _id2 = generate_stable_id("爱", "ài", "star-1")
    assert _id1 == _id2
    assert _id1.startswith("teacher-star-1-")
    assert len(_id1) == len("teacher-star-1-") + 12

    _id3 = generate_stable_id("爱", "ài", "star-2")
    assert _id3.startswith("teacher-star-2-")
    assert _id1 != _id3

    _id4 = generate_stable_id(" 爱 ", "Ài", "star-1")
    assert _id4 == _id1

    _id5 = generate_stable_id("猫", "māo", "star-1")
    assert _id1 != _id5


def _test_illustration_id():
    """Verify illustration ID generation."""
    _vid = generate_stable_id("爱", "ài", "star-1")
    _iid = generate_illustration_id(_vid)
    assert _iid == f"ill-{_vid}"


def _test_duplicate_rejection():
    """Verify duplicate normalized identities reject deterministically."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _ws.append(["爱", "ài", "愛情", "☆"])
        _ws.append([" 爱 ", "Ài", "愛", "☆"])
        _path = os.path.join(_td, "dupes.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["accepted"] == 1
        assert _m["rejected"] == 2
        _dup_count = sum(1 for r in _m["rejectedRows"] if "duplicate" in r["reason"])
        assert _dup_count == 2
        assert len(_m["duplicateDiagnostics"]) == 2


def _test_output_record_shape():
    """Verify draft output record shape matches #111/#120 contract."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "shape.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["accepted"] == 1
        _batch_path = os.path.join(_out, _m["batches"][0]["filename"])
        with open(_batch_path, "r", encoding="utf-8") as _f:
            _batch = json.load(_f)
        _entry = _batch["vocabulary"][0]

        assert "id" in _entry
        assert _entry["simplified"] == "爱"
        assert _entry["simplifiedStatus"] == "authored"
        assert _entry["pinyin"] == "ài"
        assert _entry["japanese"] == "愛する"
        assert _entry["source"] == {"type": "teacher-workbook"}
        assert _entry["reviewStatus"] == "draft"

        _cur = _entry["curriculum"]
        assert _cur["sourceId"] == "teacher-core-v1"
        assert _cur["difficultyBand"] == "star-1"
        assert _cur["sourceDifficultyLabel"] == "☆"
        assert _cur["partOfSpeech"] == "noun"
        assert _cur["sourceSheet"] == "名词1"
        assert _cur["sourceRow"] == 2
        assert "illustrationRef" not in _entry


def _test_deterministic_global_sorting():
    """Verify accepted entries are globally sorted by the defined key."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "sort.xlsx")
        _wb = Workbook()
        _default = _wb.active
        _wb.remove(_default)

        _ws1 = _wb.create_sheet("动词1")
        _ws1.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws1.append(["吃", "chī", "食べる", "☆"])

        _ws2 = _wb.create_sheet("名词1")
        _ws2.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws2.append(["爱", "ài", "愛する", "☆"])

        _ws3 = _wb.create_sheet("名词2")
        _ws3.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws3.append(["学校", "xuéxiào", "学校", "☆☆"])
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["accepted"] == 3

        _batch_path = os.path.join(_out, _m["batches"][0]["filename"])
        with open(_batch_path, "r", encoding="utf-8") as _f:
            _batch = json.load(_f)
        _entries = _batch["vocabulary"]
        assert _entries[0]["curriculum"]["sourceSheet"] == "名词1"
        assert _entries[1]["curriculum"]["sourceSheet"] == "动词1"
        assert _entries[2]["curriculum"]["sourceSheet"] == "名词2"


# ─── Phase 3 tests: Batches and manifest ────────────────────────────────────

def _test_batching():
    """Verify 50-record batching with exact filenames."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "batch.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        for _i in range(1, 52):
            _ws.append([f"字{_i}", f"zì{_i}", "テスト", "☆"])
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["accepted"] == 51
        assert _m["batchCount"] == 2
        assert _m["batches"][0]["filename"] == "teacher-vocabulary-batch-01.json"
        assert _m["batches"][1]["filename"] == "teacher-vocabulary-batch-02.json"
        assert _m["batches"][0]["entryCount"] == 50
        assert _m["batches"][1]["entryCount"] == 1


def _test_manifest_counts():
    """Verify manifest count fields are consistent."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        # All accepted
        _path = os.path.join(_td, "counts.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _ws.append(["猫", "māo", "猫", "☆"])
        _ws.append(["狗", "gǒu", "犬", "☆"])
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        assert _m["totalRows"] == 3
        assert _m["accepted"] == 3
        assert _m["rejected"] == 0
        _check_manifest_counts(_m)

        # With some rejected
        _path2 = os.path.join(_td, "counts_reject.xlsx")
        _wb2 = Workbook()
        _ws2 = _wb2.active
        _ws2.title = "名词1"
        _ws2.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws2.append(["爱", "ài", "愛する", "☆"])
        _ws2.append(["", "shénme", "何", "☆"])
        _wb2.save(_path2)
        _out2 = os.path.join(_td, "out2")
        _m2 = import_xlsx(_path2, _out2)
        assert _m2["totalRows"] == 2
        assert _m2["accepted"] == 1
        assert _m2["rejected"] == 1
        _check_manifest_counts(_m2)


def _test_deterministic_ordering():
    """Verify accepted items and rejected rows are deterministically ordered."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "order.xlsx")
        _wb = Workbook()
        _default = _wb.active
        _wb.remove(_default)
        _ws1 = _wb.create_sheet("名词1")
        _ws1.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws1.append(["爱", "ài", "愛する", "☆"])
        _ws1.append(["狗", "gǒu", "犬", "☆"])
        _ws1.append(["", "shénme", "何", "☆"])
        _ws2 = _wb.create_sheet("动词1")
        _ws2.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws2.append(["吃", "chī", "食べる", "☆"])
        _ws2.append(["", "hē", "飲む", "☆"])
        _wb.save(_path)

        _out_a = os.path.join(_td, "out-a")
        _out_b = os.path.join(_td, "out-b")
        _ma = import_xlsx(_path, _out_a)
        _mb = import_xlsx(_path, _out_b)

        with open(os.path.join(_out_a, "manifest.json"), "r", encoding="utf-8") as _f:
            _ma_data = json.load(_f)
        with open(os.path.join(_out_b, "manifest.json"), "r", encoding="utf-8") as _f:
            _mb_data = json.load(_f)

        assert _ma_data["sourceChecksumSha256"] == _mb_data["sourceChecksumSha256"]
        assert _ma_data == _mb_data


def _test_output_mutation_safety():
    """Verify fatal workbook errors stop before any output file mutation."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _default = _wb.active
        _wb.remove(_default)
        _ws = _wb.create_sheet("名词1")
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _wb.create_sheet("Unknown").append(["data"])
        _path = os.path.join(_td, "fatal_error.xlsx")
        _wb.save(_path)
        _out = os.path.join(_td, "out")
        try:
            import_xlsx(_path, _out)
        except ValueError:
            pass
        assert not os.path.exists(_out), \
            f"Output directory {_out} must not be created on fatal error"


def _test_nonempty_output_dir_fails():
    """Verify non-empty output directory fails with existing files untouched."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _xlsx = os.path.join(_td, "x.xlsx")
        _wb.save(_xlsx)

        # Create output dir with an existing file
        _out = os.path.join(_td, "out")
        os.makedirs(_out)
        _existing = os.path.join(_out, "existing.txt")
        with open(_existing, "w") as _f:
            _f.write("original")

        try:
            import_xlsx(_xlsx, _out)
            raise AssertionError("expected ValueError for non-empty output dir")
        except ValueError as e:
            assert "non-empty" in str(e).lower()
            # Existing file must be untouched
            with open(_existing, "r") as _f:
                assert _f.read() == "original", "Existing file was modified"


def _test_empty_output_dir_ok():
    """Verify a pre-existing empty output directory is allowed."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度"])
        _ws.append(["爱", "ài", "愛する", "☆"])
        _xlsx = os.path.join(_td, "x.xlsx")
        _wb.save(_xlsx)

        _out = os.path.join(_td, "out")
        os.makedirs(_out)
        _m = import_xlsx(_xlsx, _out)  # should work
        assert _m["accepted"] == 1


def _test_only_ignored_columns_row():
    """Verify a row with content only in ignored/unknown columns is a
    candidate and rejected for missing required fields (not silently skipped)."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度", "造词/造句", "备注", "ExtraCol"])
        # Row with content only in ignored and unknown columns
        _ws.append(["", "", "", "", "造句内容", "備考", "extra"])
        # Valid row to confirm import runs
        _ws.append(["爱", "ài", "愛する", "☆", "", "", ""])
        _xlsx = os.path.join(_td, "x.xlsx")
        _wb.save(_xlsx)
        _out = os.path.join(_td, "out")
        _m = import_xlsx(_xlsx, _out)

        # Row 2 has content in ignored/unknown columns → not an empty row →
        # it's a candidate. Missing required text → rejected.
        assert _m["totalRows"] == 2
        assert _m["rejected"] == 1
        assert _m["accepted"] == 1
        # Verify the rejected row is row 2 with missing simplified
        _rej = _m["rejectedRows"][0]
        assert _rej["sheet"] == "名词1"
        assert _rej["row"] == 2


def _test_ignored_column_formula():
    """Verify formula in an ignored column is not copied and does not
    trigger required-formula rejection."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _path = os.path.join(_td, "x.xlsx")
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        _ws.append(["单词", "拼音", "日语翻译", "难易度", "造词/造句"])
        _ws.append(["爱", "ài", "愛する", "☆", ""])
        _wb.save(_path)

        # Inject formula into the ignored column (造词/造句, column E = 5)
        _inject_formula(_path, "名词1", 2, 5, "=B2&C2")

        _out = os.path.join(_td, "out")
        _m = import_xlsx(_path, _out)
        # Formula in ignored column must not cause rejection
        assert _m["accepted"] == 1
        assert _m["rejected"] == 0
        # Verify output record does NOT contain the ignored-column formula
        _bp = os.path.join(_out, _m["batches"][0]["filename"])
        with open(_bp, "r", encoding="utf-8") as _f:
            _batch = json.load(_f)
        _entry = _batch["vocabulary"][0]
        assert "造词" not in str(_entry)
        assert "B2" not in str(_entry)


def _test_duplicate_header_fatal():
    """Verify duplicate required header fails before any output mutation."""
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wb = Workbook()
        _ws = _wb.active
        _ws.title = "名词1"
        # Two 单词 columns (column A and column E)
        _ws.append(["单词", "拼音", "日语翻译", "难易度", "单词"])
        _ws.append(["爱", "ài", "愛する", "☆", "extra_word"])
        _xlsx = os.path.join(_td, "dup_header.xlsx")
        _wb.save(_xlsx)

        _out = os.path.join(_td, "out")
        try:
            import_xlsx(_xlsx, _out)
            raise AssertionError("expected ValueError for duplicate header")
        except ValueError as e:
            assert "Duplicate" in str(e) or "duplicate" in str(e).lower()
            assert "单词" in str(e)
            assert "名词1" in str(e)
            # No output directory created
            assert not os.path.exists(_out)


# ─── Run all tests ───────────────────────────────────────────────────────────

def run_tests():
    """Execute all importer self-tests.  Returns 0 on success, 1 on failure."""
    errors = 0

    try:
        # Phase 1: Workbook parsing
        print("── Phase 1: Workbook parsing ──")
        tests_p1 = [
            ("Sheet classification", _test_sheet_classification),
            ("Exact headers", _test_exact_headers),
            ("Ignored headers", _test_ignored_headers),
            ("Empty/partial rows", _test_empty_partial_rows),
            ("Required formula rejection", _test_required_formula_rejection),
            ("Unknown column reporting", _test_unknown_column_reporting),
            ("Unknown sheet fatal", _test_unknown_sheet_fatal),
            ("Duplicate header fatal", _test_duplicate_header_fatal),
        ]
        for _label, _test_fn in tests_p1:
            print(f"  {_label} ... ", end="")
            _test_fn()
            print("PASS")

        # Phase 2: Records and deterministic identity
        print("── Phase 2: Records and deterministic identity ──")
        tests_p2 = [
            ("Normalization", _test_normalization),
            ("Vocabulary ID", _test_vocabulary_id),
            ("Illustration ID", _test_illustration_id),
            ("Duplicate rejection", _test_duplicate_rejection),
            ("Output record shape", _test_output_record_shape),
            ("Deterministic global sorting", _test_deterministic_global_sorting),
        ]
        for _label, _test_fn in tests_p2:
            print(f"  {_label} ... ", end="")
            _test_fn()
            print("PASS")

        # Phase 3: Batches and manifest
        print("── Phase 3: Batches and manifest ──")
        tests_p3 = [
            ("50-record batching", _test_batching),
            ("Manifest counts", _test_manifest_counts),
            ("Deterministic ordering", _test_deterministic_ordering),
            ("Output mutation safety", _test_output_mutation_safety),
            ("Non-empty output dir fails", _test_nonempty_output_dir_fails),
            ("Empty output dir ok", _test_empty_output_dir_ok),
            ("Only ignored columns row", _test_only_ignored_columns_row),
            ("Ignored column formula", _test_ignored_column_formula),
        ]
        for _label, _test_fn in tests_p3:
            print(f"  {_label} ... ", end="")
            _test_fn()
            print("PASS")

        # Phase 4: End-to-end determinism
        print("── Phase 4: End-to-end determinism ──")
        with tempfile.TemporaryDirectory() as _td:
            _fixture = os.path.join(_td, "synthetic.xlsx")
            _out_a = os.path.join(_td, "out-a")
            _out_b = os.path.join(_td, "out-b")
            os.makedirs(_out_a)
            os.makedirs(_out_b)

            _wb = _make_synthetic_wb(_fixture)
            _wb.save(_fixture)

            print("  Content validator passes ... ", end="")
            _m = import_xlsx(_fixture, _out_a)
            _check_batch_order_and_size(_out_a)
            print("PASS")

            print("  Byte-identical re-run ... ", end="")
            _m_b = import_xlsx(_fixture, _out_b)

            for _b in _m["batches"]:
                _pa = os.path.join(_out_a, _b["filename"])
                _pb = os.path.join(_out_b, _b["filename"])
                with open(_pa, "r", encoding="utf-8") as _f:
                    _da = _f.read()
                with open(_pb, "r", encoding="utf-8") as _f:
                    _db = _f.read()
                assert _da == _db, f"Byte mismatch on {_b['filename']}"

            with open(os.path.join(_out_a, "manifest.json"), "r", encoding="utf-8") as _f:
                _ma = json.load(_f)
            with open(os.path.join(_out_b, "manifest.json"), "r", encoding="utf-8") as _f:
                _mb = json.load(_f)
            assert _ma == _mb
            print("PASS")

            print("  Batch boundary 49/50/51 ... ", end="")
            for _count, _expected_batches in [(49, 1), (50, 1), (51, 2)]:
                _wb_b = Workbook()
                _ws_b = _wb_b.active
                _ws_b.title = "名词1"
                _ws_b.append(["单词", "拼音", "日语翻译", "难易度"])
                for _j in range(1, _count + 1):
                    _ws_b.append([f"字{_j}", f"zì{_j}", "テスト", "☆"])
                _bd = os.path.join(_td, f"boundary-{_count}")
                _bf = os.path.join(_td, f"boundary-{_count}.xlsx")
                _wb_b.save(_bf)
                _bm = import_xlsx(_bf, _bd)
                assert _bm["batchCount"] == _expected_batches
            print("PASS")

            print("  Manifest count consistency ... ", end="")
            _check_manifest_counts(_m)
            print("PASS")

    except Exception as e:
        print(f"\nFAILED: {e}", file=sys.stderr)
        errors = 1

    return errors


if __name__ == "__main__":
    main()
