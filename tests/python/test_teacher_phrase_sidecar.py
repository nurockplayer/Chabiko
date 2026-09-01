#!/usr/bin/env python3
from __future__ import annotations

import copy
import hashlib
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

from scripts.teacher_phrase_sidecar import (
    ContractError,
    build_sidecar,
    extract_source_units,
    phrase_id,
    refresh_sidecar,
    source_revision,
    validate_sidecar,
)


class TeacherPhraseSidecarTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.workbook_path = self.root / "teacher.xlsx"

    def tearDown(self) -> None:
        self.temp.cleanup()

    def write_workbook(self, raw_cell: object) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "名词1"
        sheet.append(["单词", "造词/造句"])
        sheet.append(["大家", raw_cell])
        workbook.save(self.workbook_path)

    def manifest(self, *, example: str = "大家好 大家请听") -> dict:
        return {
            "schemaVersion": 1,
            "source": {
                "workbookSha256": hashlib.sha256(self.workbook_path.read_bytes()).hexdigest(),
            },
            "rows": [
                {
                    "learnerId": "teacher-learner-test",
                    "simplified": "大家",
                    "sourceSheet": "名词1",
                    "sourceRow": 2,
                    "example": example,
                }
            ],
        }

    def test_build_preserves_exact_raw_lf_before_normalization(self) -> None:
        raw_cell = "  大家好\n大家请听  "
        self.write_workbook(raw_cell)

        sidecar = build_sidecar(self.manifest(), self.workbook_path)

        self.assertEqual(len(sidecar["records"]), 1)
        record = sidecar["records"][0]
        self.assertEqual(record["source"]["rawCell"], raw_cell)
        self.assertEqual(
            record["source"]["rawCellSha256"],
            hashlib.sha256(raw_cell.encode("utf-8")).hexdigest(),
        )
        self.assertEqual(record["source"]["segmentation"], "raw-lf")
        self.assertEqual(
            [phrase["simplified"] for phrase in record["teacherPhrases"]],
            ["大家好", "大家请听"],
        )
        self.assertEqual(
            [raw_cell[phrase["sourceRange"]["start"]:phrase["sourceRange"]["end"]]
             for phrase in record["teacherPhrases"]],
            ["大家好", "大家请听"],
        )

    def test_spaces_and_punctuation_never_create_draft_boundaries(self) -> None:
        raw_cell = "王先生很帅 先生，您好。"
        self.write_workbook(raw_cell)

        sidecar = build_sidecar(self.manifest(example=raw_cell), self.workbook_path)

        record = sidecar["records"][0]
        self.assertEqual(record["source"]["segmentation"], "review-required")
        self.assertEqual(record["source"]["segmentationReason"], "no-raw-lf")
        self.assertEqual(
            [phrase["simplified"] for phrase in record["teacherPhrases"]],
            [raw_cell],
        )

    def test_duplicate_source_units_require_explicit_stable_discriminators(self) -> None:
        raw_cell = "大家好\n大家好"
        self.write_workbook(raw_cell)
        manifest = self.manifest(example="大家好 大家好")

        with self.assertRaisesRegex(ContractError, "explicit duplicate discriminator"):
            build_sidecar(manifest, self.workbook_path)

        sidecar = build_sidecar(
            manifest,
            self.workbook_path,
            duplicate_discriminators={
                "teacher-learner-test": {0: "first-use", 1: "second-use"},
            },
        )
        phrases = sidecar["records"][0]["teacherPhrases"]
        self.assertEqual(
            [phrase["duplicateDiscriminator"] for phrase in phrases],
            ["first-use", "second-use"],
        )
        self.assertEqual(len({phrase["phraseId"] for phrase in phrases}), 2)

    def test_validation_rejects_wrong_learner_manifest_base(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)
        validate_sidecar(sidecar, manifest, self.workbook_path)

        drifted = copy.deepcopy(sidecar)
        drifted["base"]["learnerManifestSemanticSha256"] = "0" * 64
        with self.assertRaisesRegex(ContractError, "learner manifest semantic digest"):
            validate_sidecar(drifted, manifest, self.workbook_path)

    def test_validation_rejects_unknown_and_duplicate_learner_ids(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        unknown = copy.deepcopy(sidecar)
        unknown["records"][0]["learnerId"] = "teacher-learner-unknown"
        with self.assertRaisesRegex(ContractError, "unknown learner ID"):
            validate_sidecar(unknown, manifest, self.workbook_path)

        duplicate = copy.deepcopy(sidecar)
        duplicate["records"].append(copy.deepcopy(duplicate["records"][0]))
        with self.assertRaisesRegex(ContractError, "duplicate learner ID"):
            validate_sidecar(duplicate, manifest, self.workbook_path)

    def test_validation_rejects_source_coordinate_raw_cell_and_revision_drift(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        coordinate_drift = copy.deepcopy(sidecar)
        coordinate_drift["records"][0]["source"]["row"] = 3
        with self.assertRaisesRegex(ContractError, "source coordinate"):
            validate_sidecar(coordinate_drift, manifest, self.workbook_path)

        raw_drift = copy.deepcopy(sidecar)
        raw_drift["records"][0]["source"]["rawCell"] = "大家好！"
        with self.assertRaisesRegex(ContractError, "rawCell"):
            validate_sidecar(raw_drift, manifest, self.workbook_path)

        revision_drift = copy.deepcopy(sidecar)
        revision_drift["records"][0]["source"]["sourceRevision"] = "0" * 64
        with self.assertRaisesRegex(ContractError, "sourceRevision"):
            validate_sidecar(revision_drift, manifest, self.workbook_path)

    def test_validation_rejects_phrase_text_range_and_identity_drift(self) -> None:
        raw_cell = "大家好\n大家请听"
        self.write_workbook(raw_cell)
        manifest = self.manifest(example="大家好 大家请听")
        sidecar = build_sidecar(manifest, self.workbook_path)

        text_drift = copy.deepcopy(sidecar)
        text_drift["records"][0]["teacherPhrases"][0]["simplified"] = "大家好！"
        with self.assertRaisesRegex(ContractError, "source unit"):
            validate_sidecar(text_drift, manifest, self.workbook_path)

        range_drift = copy.deepcopy(sidecar)
        range_drift["records"][0]["teacherPhrases"][0]["sourceRange"]["start"] += 1
        with self.assertRaisesRegex(ContractError, "sourceRange"):
            validate_sidecar(range_drift, manifest, self.workbook_path)

        identity_drift = copy.deepcopy(sidecar)
        identity_drift["records"][0]["teacherPhrases"][0]["phraseId"] = "teacher-phrase-v1-" + "0" * 64
        with self.assertRaisesRegex(ContractError, "phraseId"):
            validate_sidecar(identity_drift, manifest, self.workbook_path)

    def test_field_provenance_is_per_phrase_field_with_explicit_rights_evidence(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        candidate = copy.deepcopy(sidecar)
        phrase = candidate["records"][0]["teacherPhrases"][0]
        phrase["pinyin"] = "dàjiā hǎo"
        phrase["fieldProvenance"]["pinyin"] = {
            "provenance": "generated",
            "sourceRef": "internal-model:fixture",
            "rightsRef": "repo-policy:no-external-source",
        }
        validate_sidecar(candidate, manifest, self.workbook_path)

        missing_provenance = copy.deepcopy(candidate)
        del missing_provenance["records"][0]["teacherPhrases"][0]["fieldProvenance"]["pinyin"]
        with self.assertRaisesRegex(ContractError, "fieldProvenance keys"):
            validate_sidecar(missing_provenance, manifest, self.workbook_path)

        missing_rights = copy.deepcopy(candidate)
        del missing_rights["records"][0]["teacherPhrases"][0]["fieldProvenance"]["pinyin"]["rightsRef"]
        with self.assertRaisesRegex(ContractError, "rightsRef"):
            validate_sidecar(missing_rights, manifest, self.workbook_path)

    def test_build_rejects_duplicate_learner_and_source_identities(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        manifest["rows"].append(copy.deepcopy(manifest["rows"][0]))

        with self.assertRaisesRegex(ContractError, "duplicate learner ID"):
            build_sidecar(manifest, self.workbook_path)

        manifest["rows"][1]["learnerId"] = "teacher-learner-other"
        with self.assertRaisesRegex(ContractError, "duplicate source coordinate"):
            build_sidecar(manifest, self.workbook_path)

    def test_refresh_preserves_candidate_fields_and_fails_on_source_drift(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        candidate = build_sidecar(manifest, self.workbook_path)
        phrase = candidate["records"][0]["teacherPhrases"][0]
        phrase["japanese"] = "みなさん、こんにちは"
        phrase["fieldProvenance"]["japanese"] = {
            "provenance": "authored",
            "sourceRef": "teacher-editor:fixture",
            "rightsRef": "teacher-owned:fixture",
        }
        source_revision_before = candidate["records"][0]["source"]["sourceRevision"]

        refreshed = refresh_sidecar(manifest, self.workbook_path, existing=candidate)
        self.assertEqual(refreshed, candidate)
        self.assertEqual(
            refreshed["records"][0]["source"]["sourceRevision"],
            source_revision_before,
        )

        self.write_workbook("大家好！")
        drifted_manifest = self.manifest(example="大家好！")
        before = copy.deepcopy(candidate)
        with self.assertRaisesRegex(ContractError, "digest mismatch"):
            refresh_sidecar(drifted_manifest, self.workbook_path, existing=candidate)
        self.assertEqual(candidate, before)

    def test_canonical_cli_writes_checks_and_preserves_dirty_neighbors(self) -> None:
        self.write_workbook("大家好")
        manifest_path = self.root / "learner-manifest.json"
        manifest_path.write_text(
            json.dumps(self.manifest(example="大家好"), ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        output_path = self.root / "authoring" / "teacher-phrases.json"
        output_path.parent.mkdir()
        dirty_neighbor = output_path.parent / "developer-notes.txt"
        dirty_neighbor.write_text("keep me", encoding="utf-8")
        command = [
            sys.executable,
            str(REPO_ROOT / "scripts/build-teacher-phrase-sidecar.py"),
            "--manifest",
            str(manifest_path),
            "--workbook",
            str(self.workbook_path),
            "--output",
            str(output_path),
        ]

        write_result = subprocess.run(
            [*command, "--write"],
            capture_output=True,
            check=False,
            text=True,
        )
        self.assertEqual(write_result.returncode, 0, write_result.stderr)
        first_bytes = output_path.read_bytes()

        check_result = subprocess.run(
            [*command, "--check"],
            capture_output=True,
            check=False,
            text=True,
        )
        self.assertEqual(check_result.returncode, 0, check_result.stderr)
        self.assertEqual(output_path.read_bytes(), first_bytes)
        self.assertEqual(dirty_neighbor.read_text(encoding="utf-8"), "keep me")

        self.write_workbook("大家好！")
        manifest_path.write_text(
            json.dumps(self.manifest(example="大家好！"), ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        failed_refresh = subprocess.run(
            [*command, "--write"],
            capture_output=True,
            check=False,
            text=True,
        )
        self.assertEqual(failed_refresh.returncode, 1)
        self.assertEqual(output_path.read_bytes(), first_bytes)
        self.assertEqual(dirty_neighbor.read_text(encoding="utf-8"), "keep me")

    def test_validation_rejects_mutable_review_claims_and_unknown_fields(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        mutable_review_claim = copy.deepcopy(sidecar)
        mutable_review_claim["records"][0]["reviewStatus"] = "reviewed"
        with self.assertRaisesRegex(ContractError, "record keys"):
            validate_sidecar(mutable_review_claim, manifest, self.workbook_path)

        unknown_root = copy.deepcopy(sidecar)
        unknown_root["promoted"] = True
        with self.assertRaisesRegex(ContractError, "root keys"):
            validate_sidecar(unknown_root, manifest, self.workbook_path)

    def test_unsupported_source_shapes_are_review_required_or_fail_closed(self) -> None:
        raw_crlf = "大家好\r\n大家请听"
        segmentation, reason, ranges = extract_source_units(raw_crlf)
        self.assertEqual(segmentation, "review-required")
        self.assertEqual(reason, "unsupported-carriage-return")
        self.assertEqual(ranges, [(0, len(raw_crlf))])

        for unsupported, example in (
            (42, "42"),
            ("   ", ""),
            ("=1+1", "=1+1"),
            ("#N/A", "#N/A"),
        ):
            with self.subTest(unsupported=unsupported):
                self.write_workbook(unsupported)
                with self.assertRaisesRegex(ContractError, "plain text|whitespace-only"):
                    build_sidecar(
                        self.manifest(example=example),
                        self.workbook_path,
                    )

    def test_duplicate_source_headers_fail_closed_in_build_and_validation(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        workbook = load_workbook(self.workbook_path)
        worksheet = workbook["名词1"]
        worksheet["C1"] = "造词/造句"
        worksheet["C2"] = "大家好"
        workbook.save(self.workbook_path)
        workbook.close()

        duplicate_manifest = self.manifest(example="大家好")
        with self.assertRaisesRegex(ContractError, "duplicate source header"):
            build_sidecar(duplicate_manifest, self.workbook_path)

        workbook_digest = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        sidecar["base"]["workbookSha256"] = workbook_digest
        sidecar["records"][0]["teacherPhrases"][0]["fieldProvenance"]["simplified"][
            "sourceRef"
        ] = f"teacher-workbook:sha256:{workbook_digest}#名词1:2:造词/造句"
        with self.assertRaisesRegex(ContractError, "duplicate source header"):
            validate_sidecar(sidecar, duplicate_manifest, self.workbook_path)

    def test_phrase_identity_is_position_independent_and_semantic_changes_invalidate_it(self) -> None:
        learner_id = "teacher-learner-test"
        composed = phrase_id(learner_id, "  café  ")
        decomposed = phrase_id(learner_id, "cafe\u0301")
        changed = phrase_id(learner_id, "café！")

        self.assertEqual(composed, decomposed)
        self.assertNotEqual(composed, changed)
        self.assertRegex(composed, r"^teacher-phrase-v1-[0-9a-f]{64}$")

        raw_digest = hashlib.sha256("大家好".encode("utf-8")).hexdigest()
        revision = source_revision(learner_id, "名词1", 2, raw_digest)
        self.assertNotEqual(revision, source_revision(learner_id, "名词1", 3, raw_digest))
        self.assertNotEqual(
            revision,
            source_revision(
                learner_id,
                "名词1",
                2,
                hashlib.sha256("大家好！".encode("utf-8")).hexdigest(),
            ),
        )

    def test_validation_rejects_wrong_workbook_even_when_the_bound_cell_is_unchanged(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)

        workbook = load_workbook(self.workbook_path)
        workbook["名词1"]["A3"] = "unrelated workbook drift"
        workbook.save(self.workbook_path)
        workbook.close()

        with self.assertRaisesRegex(ContractError, "workbook digest"):
            validate_sidecar(sidecar, manifest, self.workbook_path)

    def test_build_rejects_a_workbook_outside_the_manifest_source_contract(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        manifest["source"]["workbookSha256"] = "0" * 64

        with self.assertRaisesRegex(ContractError, "manifest workbook digest"):
            build_sidecar(manifest, self.workbook_path)

    def test_build_rejects_unused_or_misapplied_duplicate_discriminators(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")

        with self.assertRaisesRegex(ContractError, "unused duplicate discriminator"):
            build_sidecar(
                manifest,
                self.workbook_path,
                duplicate_discriminators={"teacher-learner-test": {0: "not-a-duplicate"}},
            )

    def test_source_simplified_provenance_stays_bound_to_workbook_rights(self) -> None:
        self.write_workbook("大家好")
        manifest = self.manifest(example="大家好")
        sidecar = build_sidecar(manifest, self.workbook_path)
        evidence = sidecar["records"][0]["teacherPhrases"][0]["fieldProvenance"]["simplified"]

        wrong_provenance = copy.deepcopy(sidecar)
        wrong_provenance["records"][0]["teacherPhrases"][0]["fieldProvenance"]["simplified"] = {
            **evidence,
            "provenance": "generated",
        }
        with self.assertRaisesRegex(ContractError, "simplified provenance"):
            validate_sidecar(wrong_provenance, manifest, self.workbook_path)

        wrong_rights = copy.deepcopy(sidecar)
        wrong_rights["records"][0]["teacherPhrases"][0]["fieldProvenance"]["simplified"] = {
            **evidence,
            "rightsRef": "unrelated-rights-record",
        }
        with self.assertRaisesRegex(ContractError, "simplified provenance"):
            validate_sidecar(wrong_rights, manifest, self.workbook_path)


if __name__ == "__main__":
    unittest.main()
